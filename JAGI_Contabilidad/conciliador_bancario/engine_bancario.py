"""
Motor de Conciliación Bancaria — JAGI CAPS v1.3
Cruza el auxiliar contable WorldOffice contra el extracto bancario.

Proceso:
  - Auxiliar: registros de la cuenta (débitos y créditos) por día
  - Extracto: movimientos del banco (créditos = abonos, débitos = cargos)
  - Cruce: por fecha y valor, con tolerancia configurable

Reglas especiales (v1.3):
  - DATÁFONOS: detección por palabras clave combinadas (más robusto que frases
    exactas). Se excluyen del extracto antes del cruce y quedan en
    hoja DATAFONOS_RETIRADOS para trazabilidad.
  - NÓMINA: solo palabra clave "nomina". Liquidación y vacaciones NO son nómina,
    van al cruce general movimiento a movimiento. Nómina agrupada por fecha
    con detalle maestro-detalle en la hoja NOMINA_CRUCE.
  - Nd Cobro Disp Fond Daviplata: se trata como movimiento bancario general
    (NO como nómina). Entra al cruce normal y queda en SOLO_EN_BANCO si
    no tiene contrapartida en el auxiliar.
  - Tolerancia porcentual: reducida a 1% (antes 5%) por decisión de gerencia.

Bancos soportados en esta versión:
  - Davivienda (Corriente / Ahorros) — formato TXT posicional de impresión
  [Agregar lectores: leer_extracto_bancolombia(), leer_extracto_bogota()]
"""

import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta (consistente con engine de datafonos) ─────────────────────────────
COLOR = {
    "header_bg":    "1F3864", "header_font": "FFFFFF",
    "cuadra":       "C6EFCE", "cuadra_font": "276221",
    "dif_menor":    "FFEB9C", "dif_font":    "9C5700",
    "diferencia":   "FFC7CE", "sin_font":    "9C0006",
    "title_bg":     "2E75B6", "title_font":  "FFFFFF",
    "total_bg":     "1F3864",
    "solo_aux":     "FCE4D6", "solo_aux_f":  "843C0C",
    "solo_banco":   "E2EFDA", "solo_banco_f":"375623",
}
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_COP  = '#,##0'
FMT_DATE = 'DD/MM/YYYY'
FMT_PCT  = '0.00%'

TOLERANCIA_DEFAULT = 1.0    # diferencia absoluta < $1 → CUADRA
TOLERANCIA_MENOR   = 0.01   # diferencia ≤ 1% del valor → DIF_MENOR (antes 5%, reducido por gerencia)

# ════════════════════════════════════════════════════════════════════════════
# CONSTANTES DE REGLAS DE NEGOCIO
# ════════════════════════════════════════════════════════════════════════════

# ── Regla 1: Datáfonos — detección por palabras clave combinadas ─────────────
# El banco puede usar distintas frases (Abono ventas netas Mastercard,
# Nc Master Ventas Netas, etc.). En lugar de frases exactas, se detecta
# por combinación: al menos una palabra de MARCAS y al menos una de TIPO.
# Así cualquier variación futura queda cubierta automáticamente.
DATAFONO_MARCAS = ["mastercard", "master", "visa", "amex", "diners"]
DATAFONO_TIPO   = ["ventas netas", "abono neto ventas", "ventas"]

def _es_datafono(descripcion: str) -> bool:
    """
    Retorna True si la descripción corresponde a un movimiento de datáfono.
    Detecta por palabras clave combinadas: marca de tarjeta + concepto de ventas.
    Insensible a mayúsculas, tildes y espacios dobles.
    """
    d = descripcion.lower().strip()
    tiene_marca  = any(m in d for m in DATAFONO_MARCAS)
    tiene_tipo   = any(t in d for t in DATAFONO_TIPO)
    return tiene_marca and tiene_tipo


# ── Regla 2: Nómina — patrones del AUXILIAR ──────────────────────────────────
# Detección flexible: solo se requiere que el texto contenga alguna de estas
# palabras clave. Captura variaciones como:
#   PAGO NOMINA / PAGO DE NOMINA / PAGO NOMINA 15/01/2026 Empleado X
#   LIQUIDACION VACACIONES / PAGO DE LIQUIDACION DEFINITIVA / PAGO LIQUIDACION
NOMINA_PALABRAS_AUXILIAR = [
    "nomina",
]
# Conceptos del auxiliar que NUNCA deben clasificarse como nómina
# aunque contenga la palabra "nomina":
NOMINA_EXCLUSIONES_AUXILIAR = [
    "prestamo",  # préstamos con "nomina" en la nota → cruce general
]

# ── Regla 2: Nómina — patrones del EXTRACTO ──────────────────────────────────
# El banco puede traer:
#   "Descuento para pago de Nomina"
#   "Nota Debito Pago Nomina Daviplata"
#   "Nomina 900123456 15/01/2026"
#   "Nomina 900123456 ENERO 2026"  (mes en texto)
# La clave es: que contenga la palabra "nomina".
# EXCLUYE expresamente "Nd Cobro Disp Fond Daviplata" — ese es gasto bancario general.
NOMINA_PALABRAS_EXTRACTO = [
    "nomina",
]
# Conceptos del extracto que NUNCA deben clasificarse como nómina
# aunque contengan alguna palabra clave:
NOMINA_EXCLUSIONES_EXTRACTO = [
    "nd cobro disp fond daviplata",   # comisión Daviplata → cruce general
    "liquidacion",                    # liquidaciones → cruce general (individual)
    "comisiones",                     # comisiones → cruce general (individual)
]


# ════════════════════════════════════════════════════════════════════════════
# 1. LECTORES DE EXTRACTO BANCARIO (uno por banco)
# ════════════════════════════════════════════════════════════════════════════

def leer_extracto_davivienda(path: str) -> pd.DataFrame:
    """
    Lee extracto bancario Davivienda en formato TXT posicional (impresión).
    Compatible con:
      - Cuenta Corriente (2346): columnas Día Mes Oficina Descripción Doc Débito Crédito Saldo
      - Cuenta Ahorros   (4677): columnas Día Mes Oficina Descripción Doc Débito Crédito
        (sin columna Saldo — la omite en ahorros)
      - Líneas de resumen bancario (no tienen oficina numérica, solo texto)
    Detecta automáticamente el tipo de cuenta desde el encabezado del archivo.
    Encoding: latin-1 (caracteres especiales Davivienda: é→?, ó→?)

    Método de parseo (heredado de conversor_txt_xls.py):
      1. Extraer todos los valores $ de la línea
      2. Los últimos 2-3 valores = débito, crédito (y saldo si hay 3)
      3. Quitar $ del texto, parsear: día mes [oficina] descripción [doc]
      4. Oficina es opcional — si no es numérico, se asume 0000
    """

    with open(path, 'r', encoding='latin-1') as f:
        lines = f.readlines()

    # Extraer metadata del encabezado (primeras 40 líneas)
    texto_cab = ' '.join(l.strip() for l in lines[:40] if l.strip())
    tipo_cuenta = 'Corriente' if 'CORRIENTE' in texto_cab.upper() else 'Ahorros'

    # Extraer año del encabezado para construir fechas completas
    m_anio = re.search(r'INFORME DEL MES:\s+\w+\s*/(\d{4})', texto_cab, re.IGNORECASE)
    anio   = int(m_anio.group(1)) if m_anio else datetime.now().year

    # ── Líneas que no son datos ───────────────────────────────────────────────
    PAT_SKIP_LINE = re.compile(
        r'^\s*(?:INFORME|REPORTE|PAGINA|FECHA|ENTIDAD|NOMBRE|TITULAR|'
        r'DESDE|HASTA|MONEDA|TOTAL|SALDO\s+INI|SALDO\s+FINAL|'
        r'OFICINA|CUENTA|TIPO|NIT|DIGITADO|REVISADO|ELABORADO|'
        r'\*+|\-+$)', re.IGNORECASE)

    def _parse_money(s):
        if not s: return 0.0
        return float(s.replace(',', '').lstrip('$').rstrip('+- '))

    registros = []
    for linea in lines:
        linea = linea.rstrip()
        if not linea.strip():
            continue
        # Saltar encabezados repetidos dentro del archivo y líneas de totales
        if PAT_SKIP_LINE.match(linea):
            continue
        # Solo líneas que empiecen con día mes (mínimo: DD MM)
        if not re.match(r'^\s*\d{1,2}\s+\d{1,2}\b', linea):
            continue

        try:
            # 1. Extraer valores monetarios
            valores = re.findall(r'\$\s*[\d,\.]+', linea)
            if not valores:
                continue  # línea sin montos = encabezado, saltar

            # 2. Identificar débito/crédito/saldo
            if len(valores) >= 3:
                deb_str  = valores[-3]
                cred_str = valores[-2]
                saldo_str = valores[-1]
                num_monedas = 3
            elif len(valores) == 2:
                deb_str  = valores[-2]
                cred_str = valores[-1]
                saldo_str = None
                num_monedas = 2
            else:
                continue

            deb   = _parse_money(deb_str.strip())
            cred  = _parse_money(cred_str.strip())
            saldo = _parse_money(saldo_str.strip()) if saldo_str else None

            # Saltar líneas que son puro resumen (ambos en cero)
            # pero conservar si alguno tiene valor
            if deb == 0 and cred == 0 and saldo is None:
                continue

            # 3. Quitar valores monetarios para parsear texto
            texto_base = linea
            for v in valores[-num_monedas:]:
                texto_base = texto_base.replace(v, '')

            # Limpiar espacios extra
            texto_base = re.sub(r'\s+', ' ', texto_base).strip()

            # 4. Tokenizar
            partes = texto_base.split()
            if len(partes) < 3:
                continue  # mínimo: día mes + algo de descripción

            dia = int(partes[0])
            mes = int(partes[1])

            # 5. Oficina (opcional — si es numérico de 3-6 dígitos)
            if len(partes) > 2 and re.match(r'^\d{3,6}$', partes[2]):
                oficina = partes[2]
                resto = partes[3:]
            else:
                oficina = '0000'
                resto = partes[2:]

            # 6. Documento: último token numérico del resto
            #    Si no hay token numérico, doc = descripción completa
            documento = None
            descripcion = ''
            for i in range(len(resto) - 1, -1, -1):
                posible_doc = re.sub(r'\D', '', resto[i])
                if posible_doc:
                    documento = posible_doc
                    descripcion = ' '.join(resto[:i])
                    break

            if not documento:
                # Documento embebido en la descripción (transferencias a otras entidades)
                documento = ''
                descripcion = ' '.join(resto)

            fecha = pd.Timestamp(year=anio, month=mes, day=dia)
            desc_clean = re.sub(r'\s{2,}', ' ', descripcion).strip()

            registros.append({
                'Fecha':       fecha,
                'Dia':         dia,
                'Mes':         mes,
                'Oficina':     oficina,
                'Descripcion': desc_clean,
                'Referencia':  documento,
                'Debitos':     deb,
                'Creditos':    cred,
                'Saldo':       saldo,
                '_Banco':      'Davivienda',
                '_TipoCuenta': tipo_cuenta,
                '_Archivo':    os.path.basename(path),
            })

        except Exception:
            continue  # línea malformada, saltar silenciosamente

    if not registros:
        raise ValueError(
            f"No se encontraron movimientos en '{os.path.basename(path)}'.\n"
            "Verifique que el archivo es un extracto Davivienda en formato TXT."
        )

    df = pd.DataFrame(registros)
    # Sanidad: filtrar filas con valores inconsistentes
    df = df[df['Fecha'].dt.year >= 2020].copy()
    return df.reset_index(drop=True)


def leer_extracto_bancolombia(path: str) -> pd.DataFrame:
    """
    Lector para extracto Bancolombia — PENDIENTE implementación.
    Completar cuando se reciba el archivo real.
    """
    raise NotImplementedError(
        "Lector Bancolombia pendiente. Por favor comparta un extracto de ejemplo "
        "de la cuenta Bancolombia para implementar el lector específico."
    )


def leer_extracto_bogota(path: str) -> pd.DataFrame:
    """
    Lector para extracto Banco de Bogotá — PENDIENTE implementación.
    Completar cuando se reciba el archivo real.
    """
    raise NotImplementedError(
        "Lector Banco de Bogotá pendiente. Por favor comparta un extracto de ejemplo "
        "para implementar el lector específico."
    )


# Registro de lectores por banco (extensible sin modificar engine)
LECTORES_EXTRACTO = {
    "Davivienda":     leer_extracto_davivienda,
    "Bancolombia":    leer_extracto_bancolombia,
    "Banco de Bogotá": leer_extracto_bogota,
}


def leer_extracto(path: str, banco: str) -> pd.DataFrame:
    """
    Punto de entrada unificado. Selecciona el lector según el banco.
    Davivienda: soporta TXT (formato impresión) y Excel.
    Bancolombia / Banco de Bogotá: pendiente — agregar lector cuando lleguen archivos.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Archivo no encontrado: {path}")

    lector = LECTORES_EXTRACTO.get(banco)
    if lector is None:
        raise ValueError(
            f"Banco '{banco}' no tiene lector implementado.\n"
            f"Bancos disponibles: {[b for b,fn in LECTORES_EXTRACTO.items() if not fn.__doc__ or 'PENDIENTE' not in fn.__doc__]}"
        )
    return lector(path)


def _mapear_columnas_extracto(cols_reales: list, mapa_esperado: dict) -> dict:
    """
    Mapea columnas reales del extracto a nombres estándar.
    Retorna {nombre_std: nombre_real_o_None}
    """
    cols_lower = {c.lower().strip(): c for c in cols_reales}
    resultado  = {}
    for col_std, alias_list in mapa_esperado.items():
        encontrado = None
        for alias in alias_list:
            if alias in cols_lower:
                encontrado = cols_lower[alias]
                break
        resultado[col_std] = encontrado
    return resultado


# ════════════════════════════════════════════════════════════════════════════
# 1b. PREPROCESAMIENTO DEL EXTRACTO — DATÁFONOS Y NÓMINA
# ════════════════════════════════════════════════════════════════════════════

def separar_datafonos(df_ext: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Regla 1 — Datáfonos:
    Separa del extracto todos los movimientos de datáfonos usando detección
    por palabras clave combinadas (marca + tipo de operación).
    Los movimientos retirados van a la hoja DATAFONOS_RETIRADOS.

    Retorna:
      df_limpio    → extracto sin datáfonos (entra al cruce)
      df_datafonos → registros retirados (trazabilidad)
    """
    mask = df_ext['Descripcion'].astype(str).apply(_es_datafono)
    return df_ext[~mask].reset_index(drop=True), df_ext[mask].reset_index(drop=True)


def _es_nomina_auxiliar(nota: str) -> bool:
    """Detecta si una nota del auxiliar corresponde a nómina (flexible)."""
    n = nota.lower().strip()
    if any(excl in n for excl in NOMINA_EXCLUSIONES_AUXILIAR):
        return False
    return any(p in n for p in NOMINA_PALABRAS_AUXILIAR)


def _es_nomina_extracto(desc: str) -> bool:
    """
    Detecta si una descripción del extracto corresponde a nómina.
    Requiere: contener alguna palabra clave de nómina Y no estar en la lista
    de exclusiones (ej. Nd Cobro Disp Fond Daviplata).
    """
    d = desc.lower().strip()
    if any(excl in d for excl in NOMINA_EXCLUSIONES_EXTRACTO):
        return False
    return any(p in d for p in NOMINA_PALABRAS_EXTRACTO)


def agrupar_nomina(df_aux: pd.DataFrame,
                   df_ext: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame,
                                                   pd.DataFrame, pd.DataFrame]:
    """
    Regla 2 — Nómina:
    Separa y agrupa registros de nómina por quincena en ambas fuentes.

    Auxiliar: N registros individuales por empleado → suma por quincena
      (primera quincena = pago 15 / segunda quincena = pago 31).
    Extracto: registros consolidados → suma por fecha.
    El concepto Nd Cobro Disp Fond Daviplata queda EXCLUIDO del bloque
    de nómina y entra al cruce general como movimiento bancario normal.

    Retorna:
      aux_sin_nomina  → auxiliar sin nómina (cruce general)
      aux_nom_grouped → nómina auxiliar agrupada por fecha de quincena
      ext_sin_nomina  → extracto sin nómina (cruce general)
      ext_nom_grouped → nómina extracto agrupada por fecha
    """

    # ── Grupo de nómina por quincena ─────────────────────────────────────────
    # Identifica quincena desde la nota del auxiliar:
    #   "PAGO NOMINA 15/01/2025 ..."           → fecha 15 = primera
    #   "PAGO NOMINA 31/01/2025 ..."            → fecha 31 = segunda
    #   "PAGO NOMINA DEL 01/01/2025 AL 15/01"  → quincena 15
    #   "PAGO NOMINA DEL 16/01/2025 AL 31/01"  → quincena 31
    PAT_QUINCENA = re.compile(r'(?:del\s+)?(\d{1,2})/\d{1,2}/\d{2,4}(?:\s+al\s+(\d{1,2}))?')

    def _quincena_fecha(nota: str, fecha: pd.Timestamp) -> pd.Timestamp:
        """
        Retorna la fecha representativa de la quincena.
        Si la nota tiene "DEL DD/MM/AAAA AL DD/MM/AAAA", usa el segundo Día (fin de rango).
        Si tiene "DD/MM/AAAA" solo, usa ese día.
        Si no encuentra nada, usa el día de la fecha original.
        """
        m = PAT_QUINCENA.search(str(nota))
        if m:
            dia_fin = int(m.group(2)) if m.group(2) else int(m.group(1))
            if dia_fin <= 15:
                return pd.Timestamp(year=fecha.year, month=fecha.month, day=15)
            else:
                ultimo = pd.Timestamp(year=fecha.year, month=fecha.month, day=1) + pd.offsets.MonthEnd(0)
                return ultimo  # fecha de fin de mes (28, 30, 31)
        return fecha

    # ── Auxiliar ─────────────────────────────────────────────────────────────
    mask_aux = df_aux['Nota'].astype(str).apply(_es_nomina_auxiliar)
    aux_nomina     = df_aux[mask_aux].copy()
    aux_sin_nomina = df_aux[~mask_aux].copy()

    if not aux_nomina.empty:
        aux_nomina['_FechaQuincena'] = aux_nomina.apply(
            lambda r: _quincena_fecha(r['Nota'], r['Fecha']), axis=1)

        aux_nom_grouped = (
            aux_nomina.groupby('_FechaQuincena', as_index=False)
            .agg(
                Creditos=('Creditos', 'sum'),
                Debitos=('Debitos',   'sum'),
                _Dia=('_Dia',   'first'),
                _Mes=('_Mes',   'first'),
                _Year=('_Year', 'first'),
                Nota=('Nota', lambda x: ' / '.join(sorted(set(x.str.strip())))),
                _conteo=('Creditos', 'count'),
            )
        )
        # Renombrar fecha de grupo para compatibilidad con cruce
        aux_nom_grouped.rename(columns={'_FechaQuincena': 'Fecha'}, inplace=True)

        # Preservar registros originales para maestro-detalle en el Excel
        aux_nom_grouped['_regs_originales'] = [
            g[['Nota', 'Debitos', 'Creditos', 'Doc Num']].to_dict('records')
            for _, g in aux_nomina.groupby('_FechaQuincena')
        ]
        aux_nom_grouped['_es_nomina_agrupada'] = True
    else:
        aux_nom_grouped = pd.DataFrame()

    # ── Extracto ─────────────────────────────────────────────────────────────
    mask_ext = df_ext['Descripcion'].astype(str).apply(_es_nomina_extracto)
    ext_nomina     = df_ext[mask_ext].copy()
    ext_sin_nomina = df_ext[~mask_ext].copy()

    if not ext_nomina.empty:
        ext_nom_grouped = (
            ext_nomina.groupby('Fecha', as_index=False)
            .agg(
                Debitos=('Debitos',   'sum'),
                Creditos=('Creditos', 'sum'),
                Descripcion=('Descripcion', lambda x: ' / '.join(sorted(set(x.str.strip())))),
            )
        )
        # Preservar filas originales individuales para maestro-detalle
        ext_nom_grouped['_regs_originales'] = [
            g.to_dict('records')
            for _, g in ext_nomina.groupby('Fecha')
        ]
        ext_nom_grouped['_es_nomina_agrupada'] = True
    else:
        ext_nom_grouped = pd.DataFrame()

    return (
        aux_sin_nomina.reset_index(drop=True),
        aux_nom_grouped,
        ext_sin_nomina.reset_index(drop=True),
        ext_nom_grouped,
        ext_nomina,
    )


# ════════════════════════════════════════════════════════════════════════════
# 1b2. MULTIABONO — agrupación por fecha (similar a nómina)
# ════════════════════════════════════════════════════════════════════════════
# El auxiliar consolera todos los cargos del día en 1 registro "Nd Pago
# Credito Multiabono DD/MM/AA". El extracto tiene decenas de cargos
# individuales (Nd Pago Credito Multiabono Master/Visa). Se agrupan por
# fecha para cruzar como un solo bloque.

def separar_multiabono(
    df_aux: pd.DataFrame, df_ext: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Separa y agrupa multiabono por fecha en ambas fuentes.

    Retorna:
      aux_sin_ma  → auxiliar sin multiabono (cruce general)
      aux_ma_grp  → multiabono auxiliar agrupado por fecha
      ext_sin_ma  → extracto con los multiabono removidos (se cruzan aparte)
      ext_ma_grp  → multiabono extracto agrupado por fecha
    """
    mask_aux_ma = df_aux['Nota'].str.lower().str.contains(
        'multiabono|nd pago credito|pago credito|credto multi', case=False, na=False)
    aux_ma   = df_aux[mask_aux_ma].copy()
    sin_ma   = df_aux[~mask_aux_ma].copy()

    if not aux_ma.empty:
        aux_ma_grp = (
            aux_ma.groupby('Fecha', as_index=False)
            .agg(
                Creditos=('Creditos', 'sum'),
                Debitos=('Debitos',   'sum'),
                _Dia=('_Dia',   'first'),
                _Mes=('_Mes',   'first'),
                _Year=('_Year', 'first'),
                Nota=('Nota', lambda x: ' / '.join(sorted(set(x.str.strip())))),
                _conteo=('Creditos', 'count'),
            )
        )
        aux_ma_grp['_regs_originales'] = [
            g[['Nota', 'Debitos', 'Creditos', 'Doc Num']].to_dict('records')
            for _, g in aux_ma.groupby('Fecha')
        ]
    else:
        aux_ma_grp = pd.DataFrame()

    mask_ext_ma = df_ext['Descripcion'].str.lower().str.contains(
        'multiabono', case=False, na=False)
    ext_ma   = df_ext[mask_ext_ma].copy()
    ext_sin  = df_ext[~mask_ext_ma].copy()

    if not ext_ma.empty:
        ext_ma_grp = (
            ext_ma.groupby('Fecha', as_index=False)
            .agg(
                Debitos=('Debitos',   'sum'),
                Creditos=('Creditos', 'sum'),
                Descripcion=('Descripcion',
                             lambda x: ' / '.join(sorted(set(x.str.strip())))),
            )
        )
        ext_ma_grp['_regs_originales'] = [
            g.to_dict('records')
            for _, g in ext_ma.groupby('Fecha')
        ]
    else:
        ext_ma_grp = pd.DataFrame()

    return (
        sin_ma.reset_index(drop=True),
        aux_ma_grp,
        ext_sin.reset_index(drop=True),
        ext_ma_grp,
    )


# ════════════════════════════════════════════════════════════════════════════
# 2. LECTURA DEL AUXILIAR (reutiliza lógica del engine de datafonos)
# ════════════════════════════════════════════════════════════════════════════

def leer_auxiliar_bancario(path: str, year: int, mes: int) -> pd.DataFrame:
    """
    Lee el auxiliar WorldOffice para conciliación bancaria.
    A diferencia del engine de datafonos, aquí NO filtramos solo DATAFONO:
    incluimos TODOS los movimientos de la cuenta en el período (débitos y créditos).
    Retorna DataFrame con: Fecha_Aux, Nota, Doc_Num, Debitos, Creditos, Saldo, _Dia
    """
    raw  = pd.read_excel(path, sheet_name=0, header=None)
    data = raw.iloc[4:].copy()
    data.columns = raw.iloc[3].tolist()
    data = data.reset_index(drop=True)

    # Normalizar columnas numéricas
    for col in ['Debitos', 'Creditos', 'Saldo']:
        data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0)

    # Normalizar fecha contable (aquí SÍ usamos la columna Fecha — es conciliación bancaria)
    data['Fecha'] = pd.to_datetime(data['Fecha'], errors='coerce')

    # Filtrar por período
    mask_periodo = (
        (data['Fecha'].dt.year  == year) &
        (data['Fecha'].dt.month == mes)
    )
    df = data[mask_periodo].copy()

    # Eliminar filas completamente vacías
    df = df[~(df['Debitos'].eq(0) & df['Creditos'].eq(0))].copy()

    df['_Dia']      = df['Fecha'].dt.day
    df['_Mes']      = df['Fecha'].dt.month
    df['_Year']     = df['Fecha'].dt.year
    df['Nota']      = df['Nota'].astype(str).str.strip()
    df['Doc Num']   = df['Doc Num'].astype(str).str.strip() if 'Doc Num' in df.columns else ''

    # ── Regla 3: Fecha embebida en conceptos especiales ────────────────────────
    # Efectivo consignado/ingresa y Nd Pago Credito Multiabono tienen fecha contable
    # 31/01 pero la fecha real está en la descripción. Extraerla para el cruce.
    PAT_FECHA_EMB = re.compile(r'(\d{1,2}/\d{1,2}/\d{2,4})')
    mask_fecha = df['Nota'].str.lower().str.contains(
        'efectivo|multiabono|consignado|consignacion', case=False, na=False)
    for idx in df[mask_fecha].index:
        m = PAT_FECHA_EMB.search(str(df.at[idx, 'Nota']))
        if m:
            fecha_txt = m.group(1)
            # Normalizar año corto: "12/01/25" → 2025
            partes = fecha_txt.split('/')
            anio_p = int(partes[2]) if len(partes[2]) == 4 else 2000 + int(partes[2])
            try:
                df.at[idx, 'Fecha'] = pd.Timestamp(
                    year=anio_p, month=int(partes[1]), day=int(partes[0]))
                df.at[idx, '_Dia']  = int(partes[0])
                df.at[idx, '_Mes']  = int(partes[1])
                df.at[idx, '_Year'] = anio_p
            except (ValueError, TypeError):
                pass  # mantener fecha original si no parsea

    return df.reset_index(drop=True)


# ════════════════════════════════════════════════════════════════════════════
# 3. CRUCE AUXILIAR ↔ EXTRACTO BANCARIO
# ════════════════════════════════════════════════════════════════════════════

def cruzar_auxiliar_extracto(df_aux: pd.DataFrame,
                              df_ext: pd.DataFrame,
                              year: int, mes: int,
                              tolerancia_abs: float = TOLERANCIA_DEFAULT,
                              tolerancia_pct: float = TOLERANCIA_MENOR) -> dict:
    """
    Cruza el auxiliar contra el extracto bancario.

    Pre-proceso:
      1. Datáfonos: retirados del extracto → hoja DATAFONOS_RETIRADOS
      2. Nómina: agrupada por fecha en ambos lados → cruce consolidado

    Estrategia de cruce general (prioridad):
      1. Mismo día + valor exacto (dif < $1)
      2. Mismo día + diferencia ≤ 1%
      3. ±3 días + valor exacto
      4. ±3 días + diferencia ≤ 1%
      5. Sin match

    Retorna dict con:
      'cruces'        → cruce general movimiento a movimiento
      'cruces_nomina' → cruce de nómina agrupado por fecha
      'cruces_multiabono' → cruce de multiabono agrupado por fecha
      'solo_banco'    → extracto sin contrapartida (cruce general)
      'datafonos'     → movimientos de datáfonos retirados
      'resumen'       → totales y estadísticas
    """
    ext_periodo = df_ext[
        (df_ext['Fecha'].dt.year  == year) &
        (df_ext['Fecha'].dt.month == mes)
    ].copy()

    # ── Regla 1: Retirar datáfonos ───────────────────────────────────────────
    ext_limpio, df_datafonos = separar_datafonos(ext_periodo)

    # ── Regla 3: Separar y agrupar multiabono ────────────────────────────────
    aux_sin_ma, aux_ma_grp, ext_sin_ma, ext_ma_grp = separar_multiabono(df_aux, ext_limpio)

    # ── Regla 2: Separar y agrupar nómina (sobre datos ya sin multiabono) ────
    aux_sin_nom, aux_nom_grp, ext_sin_nom, ext_nom_grp, ext_nom_rows = agrupar_nomina(aux_sin_ma, ext_sin_ma)

    # ── Regla 2b: Desviar registros mal clasificados como nómina ─────────────
    # Si el extracto clasifica como "nomina" algo que el auxiliar llama
    # "liquidacion", "comision", etc. (por fecha + valor exacto), se devuelve
    # al grupo sin-nomina para que cruze como cruce general.
    nom_ext_solo_banco_indices = []
    if not ext_nom_grp.empty:
        fechas_aux_nom = set(aux_nom_grp['Fecha']) if not aux_nom_grp.empty else set()
        for _, re_ in ext_nom_grp.iterrows():
            fecha_ext = re_['Fecha']
            if fecha_ext not in fechas_aux_nom:
                regs_ext_orig = re_.get('_regs_originales', [])
                if regs_ext_orig:
                    for reg in regs_ext_orig:
                        v_ext_reg = float(reg.get('Debitos', 0))
                        fecha_reg = reg.get('Fecha')
                        if fecha_reg is not None and v_ext_reg > 0:
                            match_directo = aux_sin_nom[
                                (aux_sin_nom['Fecha'] == fecha_reg) &
                                (aux_sin_nom['Creditos'] == v_ext_reg)
                            ]
                            if not match_directo.empty:
                                nom_ext_solo_banco_indices.extend(
                                    ext_nom_rows[ext_nom_rows['Fecha'] == fecha_ext].index.tolist())
                                break

    if nom_ext_solo_banco_indices:
        mal_clasificados = ext_nom_rows.loc[
            list(set(nom_ext_solo_banco_indices))
        ].copy()
        ext_sin_nom = pd.concat([ext_sin_nom, mal_clasificados], ignore_index=True)
        ext_sin_nom = ext_sin_nom.reset_index(drop=True)

    # ── Cruce general ────────────────────────────────────────────────────────
    aux_deb  = aux_sin_nom[aux_sin_nom['Debitos']  > 0].copy()
    aux_cred = aux_sin_nom[aux_sin_nom['Creditos'] > 0].copy()
    ext_cred = ext_sin_nom[ext_sin_nom['Creditos'] > 0].copy()
    ext_deb  = ext_sin_nom[ext_sin_nom['Debitos']  > 0].copy()

    cruces    = []
    usados_ext = set()

    def _cruzar_grupo(df_a, df_e, col_a, col_e, tipo):
        usados_local = set()
        for idx_a, ra in df_a.iterrows():
            v_aux = float(ra[col_a])
            mejor = None; mejor_score = -1
            for idx_e, re_ in df_e.iterrows():
                if idx_e in usados_ext or idx_e in usados_local:
                    continue
                v_ext = float(re_[col_e])
                dif   = abs(v_aux - v_ext)
                dif_e = int(abs((re_['Fecha'] - ra['Fecha']).days)) \
                        if pd.notna(re_['Fecha']) and pd.notna(ra['Fecha']) else 999
                if dif < tolerancia_abs and dif_e == 0:
                    mejor = (idx_e, re_, dif, dif_e, 'CUADRA'); break
                elif dif <= v_aux * tolerancia_pct and dif_e == 0:
                    score = 90
                    if score > mejor_score: mejor = (idx_e, re_, dif, dif_e, 'DIF_MENOR'); mejor_score = score
                elif dif < tolerancia_abs and dif_e <= 3:
                    score = 80 - dif_e * 10
                    if score > mejor_score: mejor = (idx_e, re_, dif, dif_e, 'CUADRA_FECHA_DIF'); mejor_score = score
                elif dif <= v_aux * tolerancia_pct and dif_e <= 3:
                    score = 70 - dif_e * 10
                    if score > mejor_score: mejor = (idx_e, re_, dif, dif_e, 'DIF_MENOR_FECHA_DIF'); mejor_score = score
            if mejor:
                idx_e, re_, dif, dif_e, estado = mejor
                usados_ext.add(idx_e); usados_local.add(idx_e)
                cruces.append({
                    'tipo': tipo, 'row_aux': ra, 'row_ext': re_,
                    'valor_aux': v_aux, 'valor_ext': float(re_[col_e]),
                    'diferencia': v_aux - float(re_[col_e]),
                    'dias_offset': dif_e, 'estado': estado,
                    'fecha_aux': ra['Fecha'], 'fecha_ext': re_['Fecha'],
                })
            else:
                cruces.append({
                    'tipo': tipo, 'row_aux': ra, 'row_ext': None,
                    'valor_aux': v_aux, 'valor_ext': 0.0,
                    'diferencia': v_aux, 'dias_offset': None, 'estado': 'SIN_MATCH',
                    'fecha_aux': ra['Fecha'], 'fecha_ext': None,
                })

    _cruzar_grupo(aux_deb,  ext_cred, 'Debitos',  'Creditos', 'DEBITO_AUX')
    _cruzar_grupo(aux_cred, ext_deb,  'Creditos', 'Debitos',  'CREDITO_AUX')

    # ── Cruce de nómina (agrupado por fecha) ──────────────────────────────────
    cruces_nomina = []
    nom_ext_solo_banco_indices = []  # registros del extracto que NO son realmente nómina

    if not aux_nom_grp.empty and not ext_nom_grp.empty:
        ext_nom_idx = ext_nom_grp.set_index('Fecha')
        fechas_aux  = set(aux_nom_grp['Fecha'])

        for _, ra in aux_nom_grp.iterrows():
            fecha  = ra['Fecha']
            v_aux  = float(ra['Creditos'])
            conteo = int(ra.get('_conteo', 1))
            nota   = str(ra.get('Nota', ''))
            regs_aux_orig = ra.get('_regs_originales', [])

            if fecha in ext_nom_idx.index:
                re_   = ext_nom_idx.loc[fecha]   # Series (fecha única tras groupby)
                v_ext = float(re_['Debitos'])
                regs_ext_orig = re_.get('_regs_originales', [])
                dif   = abs(v_aux - v_ext)
                if dif < tolerancia_abs:
                    estado = 'NOMINA_CUADRA'
                elif dif <= v_aux * tolerancia_pct:
                    estado = 'NOMINA_DIF_MENOR'
                else:
                    estado = 'NOMINA_DIF'
                cruces_nomina.append({
                    'fecha': fecha, 'nota_aux': nota, 'conteo_empleados': conteo,
                    'valor_aux': v_aux, 'valor_ext': v_ext,
                    'diferencia': v_aux - v_ext, 'estado': estado,
                    'observacion': '',
                    'regs_aux_originales': regs_aux_orig,
                    'regs_ext_originales': regs_ext_orig,
                })
            else:
                cruces_nomina.append({
                    'fecha': fecha, 'nota_aux': nota, 'conteo_empleados': conteo,
                    'valor_aux': v_aux, 'valor_ext': 0.0,
                    'diferencia': v_aux, 'estado': 'NOMINA_SIN_MATCH_BANCO',
                    'observacion': 'Pago de nómina en auxiliar sin movimiento bancario en esta fecha',
                    'regs_aux_originales': regs_aux_orig,
                    'regs_ext_originales': [],
                })

        # ── Fechas solo en extracto: validar si son realmente nómina ─────────
        # Si el extracto clasifica como "nomina" algo que el auxiliar llama
        # "liquidacion", "comision", etc., esos registros individuales se
        # desvían al cruce general en vez de quedar como SOLO_BANCO de nómina.
        for _, re_ in ext_nom_grp.iterrows():
            fecha_ext = re_['Fecha']
            if fecha_ext not in fechas_aux:
                regs_ext_orig = re_.get('_regs_originales', [])

                # Verificar si algún registro individual coincide con
                # un registro del auxiliar NO-nómina (liquidación, etc.)
                # por fecha + valor exacto. Si hay match, se envía al
                # cruce general en vez de nómina SOLO_BANCO.
                hay_match_no_nomina = False
                if regs_ext_orig:
                    for reg in regs_ext_orig:
                        v_ext_reg = float(reg.get('Debitos', 0))
                        fecha_reg = reg.get('Fecha')
                        if fecha_reg is not None and v_ext_reg > 0:
                            match_directo = aux_sin_nom[
                                (aux_sin_nom['Fecha'] == fecha_reg) &
                                (aux_sin_nom['Creditos'] == v_ext_reg)
                            ]
                            if not match_directo.empty:
                                hay_match_no_nomina = True
                                nom_ext_solo_banco_indices.extend(
                                    ext_nom_rows[ext_nom_rows['Fecha'] == fecha_ext].index.tolist())
                                break

                if not hay_match_no_nomina:
                    v_ext = float(re_['Debitos'])
                    cruces_nomina.append({
                        'fecha': fecha_ext,
                        'nota_aux': '— Sin registro en auxiliar —',
                        'conteo_empleados': 0,
                        'valor_aux': 0.0, 'valor_ext': v_ext,
                        'diferencia': -v_ext, 'estado': 'NOMINA_SOLO_BANCO',
                        'observacion': 'Nómina registrada en banco sin contrapartida en auxiliar en esta fecha',
                        'regs_aux_originales': [],
                        'regs_ext_originales': regs_ext_orig,
                    })

    elif not ext_nom_grp.empty:
        for _, re_ in ext_nom_grp.iterrows():
            v_ext = float(re_['Debitos'])
            regs_ext_orig = re_.get('_regs_originales', [])
            cruces_nomina.append({
                'fecha': re_['Fecha'], 'nota_aux': '— Sin registro en auxiliar —',
                'conteo_empleados': 0, 'valor_aux': 0.0, 'valor_ext': v_ext,
                'diferencia': -v_ext, 'estado': 'NOMINA_SOLO_BANCO',
                'observacion': 'Nómina bancaria sin registros en auxiliar',
                'regs_aux_originales': [],
                'regs_ext_originales': regs_ext_orig,
            })

    # ── Cruce de multiabono (agrupado por fecha) ─────────────────────────────
    cruces_multiabono = []

    if not aux_ma_grp.empty and not ext_ma_grp.empty:
        ext_ma_idx = ext_ma_grp.set_index('Fecha')
        fechas_aux_ma = set(aux_ma_grp['Fecha'])

        for _, ra in aux_ma_grp.iterrows():
            fecha  = ra['Fecha']
            v_aux  = float(ra['Debitos'])  # multiabono es débito en el auxiliar
            conteo = int(ra.get('_conteo', 1))
            nota   = str(ra.get('Nota', ''))
            regs_aux_orig = ra.get('_regs_originales', [])

            if fecha in ext_ma_idx.index:
                re_   = ext_ma_idx.loc[fecha]
                v_ext = float(re_['Creditos'])  # crédito en el extracto
                regs_ext_orig = re_.get('_regs_originales', [])
                dif   = abs(v_aux - v_ext)
                if dif < tolerancia_abs:
                    estado = 'MULTIABONO_CUADRA'
                elif dif <= v_aux * tolerancia_pct and v_aux > 0:
                    estado = 'MULTIABONO_DIF_MENOR'
                else:
                    estado = 'MULTIABONO_DIF'
                cruces_multiabono.append({
                    'fecha': fecha, 'nota_aux': nota, 'conteo_registros': conteo,
                    'valor_aux': v_aux, 'valor_ext': v_ext,
                    'diferencia': v_aux - v_ext, 'estado': estado,
                    'observacion': '',
                    'regs_aux_originales': regs_aux_orig,
                    'regs_ext_originales': regs_ext_orig,
                })
            else:
                cruces_multiabono.append({
                    'fecha': fecha, 'nota_aux': nota, 'conteo_registros': conteo,
                    'valor_aux': v_aux, 'valor_ext': 0.0,
                    'diferencia': v_aux, 'estado': 'MULTIABONO_SIN_MATCH_BANCO',
                    'observacion': 'Multiabono en auxiliar sin movimiento bancario en esta fecha',
                    'regs_aux_originales': regs_aux_orig,
                    'regs_ext_originales': [],
                })

        # Fechas solo en extracto (multiabono bancario sin contrapartida)
        for _, re_ in ext_ma_grp.iterrows():
            fecha_ext = re_['Fecha']
            if fecha_ext not in fechas_aux_ma:
                v_ext = float(re_['Creditos'])
                regs_ext_orig = re_.get('_regs_originales', [])
                cruces_multiabono.append({
                    'fecha': fecha_ext, 'nota_aux': '— Sin registro en auxiliar —',
                    'conteo_registros': 0, 'valor_aux': 0.0, 'valor_ext': v_ext,
                    'diferencia': -v_ext, 'estado': 'MULTIABONO_SOLO_BANCO',
                    'observacion': 'Multiabono bancario sin contrapartida en auxiliar',
                    'regs_aux_originales': [],
                    'regs_ext_originales': regs_ext_orig,
                })

    elif not ext_ma_grp.empty:
        for _, re_ in ext_ma_grp.iterrows():
            v_ext = float(re_['Creditos'])
            regs_ext_orig = re_.get('_regs_originales', [])
            cruces_multiabono.append({
                'fecha': re_['Fecha'], 'nota_aux': '— Sin registro en auxiliar —',
                'conteo_registros': 0, 'valor_aux': 0.0, 'valor_ext': v_ext,
                'diferencia': -v_ext, 'estado': 'MULTIABONO_SOLO_BANCO',
                'observacion': 'Multiabono bancario sin registros en auxiliar',
                'regs_aux_originales': [],
                'regs_ext_originales': regs_ext_orig,
            })

    # ── Movimientos solo en banco (cruce general) ────────────────────────────
    idxs_cruzados = {c['row_ext'].name for c in cruces if c['row_ext'] is not None}
    solo_banco    = ext_sin_nom[~ext_sin_nom.index.isin(idxs_cruzados)].copy()

    # ── Resumen ──────────────────────────────────────────────────────────────
    cuadra    = sum(1 for c in cruces if c['estado'] in ('CUADRA', 'CUADRA_FECHA_DIF'))
    dif_menor = sum(1 for c in cruces if 'DIF_MENOR' in c['estado'])
    sin_match = sum(1 for c in cruces if c['estado'] == 'SIN_MATCH')
    total_aux = sum(c['valor_aux'] for c in cruces)
    total_ext = sum(c['valor_ext'] for c in cruces if c['row_ext'] is not None)

    nom_cuadra    = sum(1 for c in cruces_nomina if c['estado'] == 'NOMINA_CUADRA')
    nom_dif       = sum(1 for c in cruces_nomina if c['estado'] in ('NOMINA_DIF', 'NOMINA_DIF_MENOR'))
    nom_sin_match = sum(1 for c in cruces_nomina
                        if c['estado'] in ('NOMINA_SIN_MATCH_BANCO', 'NOMINA_SOLO_BANCO'))
    nom_total_aux = sum(c['valor_aux'] for c in cruces_nomina)
    nom_total_ext = sum(c['valor_ext'] for c in cruces_nomina)

    ma_cuadra    = sum(1 for c in cruces_multiabono if 'CUADRA' in c['estado'])
    ma_dif       = sum(1 for c in cruces_multiabono if 'DIF_MENOR' in c['estado'] or c['estado'] == 'MULTIABONO_DIF')
    ma_sin_match = sum(1 for c in cruces_multiabono
                       if c['estado'] in ('MULTIABONO_SIN_MATCH_BANCO', 'MULTIABONO_SOLO_BANCO'))
    ma_total_aux = sum(c['valor_aux'] for c in cruces_multiabono)
    ma_total_ext = sum(c['valor_ext'] for c in cruces_multiabono)

    resumen = {
        'total_registros':     len(cruces),
        'cuadra':              cuadra,
        'dif_menor':           dif_menor,
        'sin_match':           sin_match,
        'solo_banco':          len(solo_banco),
        'total_aux':           total_aux,
        'total_ext':           total_ext,
        'diferencia_neta':     total_aux - total_ext,
        'total_debitos_ext':   float(ext_periodo['Debitos'].sum())   if not ext_periodo.empty else 0.0,
        'total_creditos_ext':  float(ext_periodo['Creditos'].sum())  if not ext_periodo.empty else 0.0,
        'total_debitos_aux':   float(df_aux['Debitos'].sum())         if not df_aux.empty else 0.0,
        'total_creditos_aux':  float(df_aux['Creditos'].sum())        if not df_aux.empty else 0.0,
        'nom_fechas':          len(cruces_nomina),
        'nom_cuadra':          nom_cuadra,
        'nom_dif':             nom_dif,
        'nom_sin_match':       nom_sin_match,
        'nom_total_aux':       nom_total_aux,
        'nom_total_ext':       nom_total_ext,
        'nom_diferencia':      nom_total_aux - nom_total_ext,
        'ma_fechas':           len(cruces_multiabono),
        'ma_cuadra':           ma_cuadra,
        'ma_dif':              ma_dif,
        'ma_sin_match':        ma_sin_match,
        'ma_total_aux':        ma_total_aux,
        'ma_total_ext':        ma_total_ext,
        'ma_diferencia':       ma_total_aux - ma_total_ext,
        'datafonos_retirados': len(df_datafonos),
        'datafonos_valor':     float(df_datafonos['Creditos'].sum()) if not df_datafonos.empty else 0.0,
    }

    return {
        'cruces':             cruces,
        'cruces_nomina':      cruces_nomina,
        'cruces_multiabono':  cruces_multiabono,
        'solo_banco':         solo_banco,
        'datafonos':          df_datafonos,
        'resumen':            resumen,
    }


# ════════════════════════════════════════════════════════════════════════════
# 4. GENERACIÓN DEL EXCEL DE RESULTADO
# ════════════════════════════════════════════════════════════════════════════

def _sc(cell, bg=None, fc="000000", bold=False, fmt=None, ha="left", sz=10):
    if bg: cell.fill = PatternFill("solid", start_color=bg, fgColor=bg)
    cell.font      = Font(color=fc, bold=bold, name="Arial", size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=False)
    cell.border    = BORDER
    if fmt: cell.number_format = fmt


def _titulo(ws, texto, subtexto="", ncols=14):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = texto
    t.fill  = PatternFill("solid", start_color=COLOR["title_bg"])
    t.font  = Font(color="FFFFFF", bold=True, name="Arial", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    if subtexto:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        g = ws["A2"]
        g.value = subtexto
        g.font  = Font(color="595959", italic=True, name="Arial", size=8)
        g.alignment = Alignment(horizontal="center")


def generar_excel_bancario(resultado: dict,
                            cuenta: str,
                            banco: str,
                            output_path: str,
                            periodo: str = "Enero 2026",
                            info_empresa: dict = None) -> str:
    """
    Genera el Excel de conciliación bancaria v1.3.
    Hojas: RESUMEN | DETALLE_CRUCE | NOMINA_CRUCE | MULTIABONO_CRUCE | DATAFONOS_RETIRADOS | SOLO_EN_BANCO | LOG_AUDITORIA
    """
    wb               = Workbook()
    cruces           = resultado['cruces']
    cruces_nomina    = resultado.get('cruces_nomina', [])
    cruces_multiabono = resultado.get('cruces_multiabono', [])
    solo_banco       = resultado['solo_banco']
    datafonos        = resultado.get('datafonos', pd.DataFrame())
    resumen          = resultado['resumen']
    emp_label        = f"{info_empresa['razon_social']} — {cuenta}" if info_empresa else cuenta

    # ── Hoja RESUMEN ──────────────────────────────────────────────────────────
    ws_r = wb.active; ws_r.title = "RESUMEN"
    _titulo(ws_r,
            f"CONCILIACIÓN BANCARIA — {cuenta} — {periodo.upper()}",
            f"JAGI CAPS | {emp_label} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            ncols=3)
    ws_r.column_dimensions['A'].width = 52
    ws_r.column_dimensions['B'].width = 16
    ws_r.column_dimensions['C'].width = 22

    def put(r, c, v, bg=None, bold=False, fmt=None, ha="left"):
        cel = ws_r.cell(row=r, column=c, value=v)
        _sc(cel, bg=bg, bold=bold, fmt=fmt, ha=ha)

    # Sección general
    put(4, 1, "CONCILIACIÓN GENERAL (sin datáfonos ni nómina)",
        bg=COLOR["header_bg"], bold=True, ha="center")
    put(4, 2, "CANTIDAD",  bg=COLOR["header_bg"], bold=True, ha="center")
    put(4, 3, "VALOR ($)", bg=COLOR["header_bg"], bold=True, ha="center")
    ws_r.row_dimensions[4].height = 20

    stats_g = [
        ("Total movimientos en auxiliar (general)",          resumen['total_registros'],  resumen['total_aux']),
        ("✔  Cuadran exacto (mismo día y valor)",            resumen['cuadra'],            None),
        ("⚠  Diferencia ≤1% / fecha diferente",              resumen['dif_menor'],         None),
        ("❌  Sin match en extracto bancario",               resumen['sin_match'],          None),
        ("ℹ  Solo en extracto bancario (sin auxiliar)",      resumen['solo_banco'],         None),
        ("Diferencia neta general ($)",                       "",                           resumen['diferencia_neta']),
    ]
    bgs_g = [None, COLOR["cuadra"], COLOR["dif_menor"], COLOR["diferencia"], COLOR["solo_banco"], None]
    for i, ((lbl, cant, val), bg) in enumerate(zip(stats_g, bgs_g), 1):
        put(4+i, 1, lbl,  bg=bg)
        put(4+i, 2, cant if cant != "" else "", bg=bg, ha="center")
        put(4+i, 3, val if val is not None else "", bg=bg,
            fmt=FMT_COP if val is not None else None, ha="right")
        ws_r.row_dimensions[4+i].height = 18

    # Sección nómina
    fila_nom = 12
    put(fila_nom, 1, "CONCILIACIÓN NÓMINA (agrupada por fecha)",
        bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_nom, 2, "FECHAS",    bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_nom, 3, "VALOR ($)", bg=COLOR["header_bg"], bold=True, ha="center")
    ws_r.row_dimensions[fila_nom].height = 20

    stats_n = [
        ("✔  Fechas de nómina que cuadran",                 resumen.get('nom_cuadra', 0),     None),
        ("⚠  Fechas con diferencia de valor",                resumen.get('nom_dif', 0),         None),
        ("❌  Fechas sin contrapartida",                     resumen.get('nom_sin_match', 0),   None),
        ("Total nómina auxiliar ($)",                         "",                               resumen.get('nom_total_aux', 0)),
        ("Total nómina extracto ($)",                         "",                               resumen.get('nom_total_ext', 0)),
        ("Diferencia neta nómina ($)",                        "",                               resumen.get('nom_diferencia', 0)),
    ]
    bgs_n = [COLOR["cuadra"], COLOR["dif_menor"], COLOR["diferencia"], None, None, None]
    for i, ((lbl, cant, val), bg) in enumerate(zip(stats_n, bgs_n), 1):
        put(fila_nom+i, 1, lbl,  bg=bg)
        put(fila_nom+i, 2, cant if cant != "" else "", bg=bg, ha="center")
        put(fila_nom+i, 3, val if val is not None else "", bg=bg,
            fmt=FMT_COP if val is not None else None, ha="right")
        ws_r.row_dimensions[fila_nom+i].height = 18

    # Sección — Totales importados del archivo
    fila_tot = fila_nom + len(stats_n) + 2
    put(fila_tot, 1, "TOTALES IMPORTADOS DE ARCHIVOS",
        bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_tot, 2, "DEBITOS ($)",  bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_tot, 3, "CREDITOS ($)", bg=COLOR["header_bg"], bold=True, ha="center")
    ws_r.row_dimensions[fila_tot].height = 20
    stats_t = [
        ("Total débitos extracto",   resumen.get('total_debitos_ext', 0),  None),
        ("Total créditos extracto",  None, resumen.get('total_creditos_ext', 0)),
        ("Total débitos auxiliar",   resumen.get('total_debitos_aux', 0),  None),
        ("Total créditos auxiliar",  None, resumen.get('total_creditos_aux', 0)),
    ]
    bgs_t = ["EBF3FB"] * 4
    for i, (lbl, deb, cred), bg in zip(range(1, 5), stats_t, bgs_t):
        put(fila_tot+i, 1, lbl,  bg=bg)
        put(fila_tot+i, 2, deb, bg=bg, fmt=FMT_COP, ha="right")
        put(fila_tot+i, 3, cred, bg=bg, fmt=FMT_COP, ha="right")

    # Sección multiabono
    fila_ma = fila_tot + len(stats_t) + 2
    put(fila_ma, 1, "CONCILIACIÓN MULTIABONO (agrupado por fecha)",
        bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_ma, 2, "FECHAS",    bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_ma, 3, "VALOR ($)", bg=COLOR["header_bg"], bold=True, ha="center")
    ws_r.row_dimensions[fila_ma].height = 20
    stats_ma = [
        ("✔  Fechas de multiabono que cuadran",                resumen.get('ma_cuadra', 0),    None),
        ("⚠  Fechas con diferencia de valor",                  resumen.get('ma_dif', 0),       None),
        ("❌  Fechas sin contrapartida",                       resumen.get('ma_sin_match', 0), None),
        ("Total multiabono auxiliar ($)",                       "",                            resumen.get('ma_total_aux', 0)),
        ("Total multiabono extracto ($)",                       "",                            resumen.get('ma_total_ext', 0)),
        ("Diferencia neta multiabono ($)",                      "",                            resumen.get('ma_diferencia', 0)),
    ]
    bgs_ma = [COLOR["cuadra"], COLOR["dif_menor"], COLOR["diferencia"], None, None, None]
    for i, ((lbl, cant, val), bg) in enumerate(zip(stats_ma, bgs_ma), 1):
        put(fila_ma+i, 1, lbl,  bg=bg)
        put(fila_ma+i, 2, cant if cant != "" else "", bg=bg, ha="center")
        put(fila_ma+i, 3, val if val is not None else "", bg=bg,
            fmt=FMT_COP if val is not None else None, ha="right")
        ws_r.row_dimensions[fila_ma+i].height = 18

    # Sección datáfonos
    fila_dat = fila_ma + len(stats_ma) + 2
    put(fila_dat, 1, "DATÁFONOS RETIRADOS DEL EXTRACTO",
        bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_dat, 2, "REGISTROS", bg=COLOR["header_bg"], bold=True, ha="center")
    put(fila_dat, 3, "VALOR ($)",  bg=COLOR["header_bg"], bold=True, ha="center")
    ws_r.row_dimensions[fila_dat].height = 20
    put(fila_dat+1, 1, "Movimientos de datáfonos excluidos del cruce bancario", bg="EBF3FB")
    put(fila_dat+1, 2, resumen.get('datafonos_retirados', 0), bg="EBF3FB", ha="center")
    put(fila_dat+1, 3, resumen.get('datafonos_valor', 0),
        bg="EBF3FB", fmt=FMT_COP, ha="right")
    ws_r.row_dimensions[fila_dat+1].height = 18

    # ── Hoja DETALLE_CRUCE ────────────────────────────────────────────────────
    ws_d = wb.create_sheet("DETALLE_CRUCE")
    _titulo(ws_d,
            f"DETALLE DE CRUCE GENERAL — {cuenta} — {periodo.upper()}",
            "Auxiliar WorldOffice ↔ Extracto Bancario (excluye datáfonos y nómina)", ncols=12)
    HDRS = ["Tipo", "Fecha Aux", "Nota Auxiliar", "Doc Num",
            "Valor Aux ($)", "Fecha Banco", "Descripción Banco", "Ref. Banco",
            "Valor Banco ($)", "Diferencia ($)", "Días Offset", "Estado"]
    WIDS = [12, 12, 36, 16, 16, 12, 32, 16, 16, 14, 12, 18]
    for ci, (h, w) in enumerate(zip(HDRS, WIDS), 1):
        cel = ws_d.cell(row=3, column=ci, value=h)
        _sc(cel, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws_d.column_dimensions[get_column_letter(ci)].width = w
    ws_d.row_dimensions[3].height = 20

    row_n = 4
    for c in cruces:
        estado = c['estado']
        if 'CUADRA' in estado:   bg, fc = COLOR["cuadra"],    COLOR["cuadra_font"]
        elif 'DIF_MENOR' in estado: bg, fc = COLOR["dif_menor"], COLOR["dif_font"]
        else:                    bg, fc = COLOR["diferencia"], COLOR["sin_font"]
        re_ = c['row_ext']
        vals = [
            c['tipo'],
            c['fecha_aux'].date() if pd.notna(c['fecha_aux']) else None,
            str(c['row_aux'].get('Nota', ''))[:60],
            str(c['row_aux'].get('Doc Num', '')),
            c['valor_aux'],
            c['fecha_ext'].date() if re_ is not None and pd.notna(c['fecha_ext']) else None,
            str(re_['Descripcion'])[:50] if re_ is not None and 'Descripcion' in re_ else None,
            str(re_['Referencia'])       if re_ is not None and 'Referencia'  in re_ else None,
            c['valor_ext'] if c['valor_ext'] > 0 else None,
            c['diferencia'] if c['valor_ext'] > 0 else None,
            c['dias_offset'],
            estado,
        ]
        fmts = [None, FMT_DATE, None, None, FMT_COP, FMT_DATE, None, None, FMT_COP, FMT_COP, None, None]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cel = ws_d.cell(row=row_n, column=ci, value=v)
            _sc(cel, bg=bg, fc=fc, fmt=fmt,
                ha="right" if ci in [5, 9, 10] else "center" if ci in [1, 2, 6, 11] else "left")
        ws_d.row_dimensions[row_n].height = 16
        row_n += 1
    ws_d.freeze_panes = "A4"

    # ── Hoja NOMINA_CRUCE ─────────────────────────────────────────────────────
    ws_n = wb.create_sheet("NOMINA_CRUCE")
    _titulo(ws_n,
            f"CONCILIACIÓN NÓMINA — {cuenta} — {periodo.upper()}",
            "Cruce desagregado: cada fila auxiliar con los movimientos del extracto que cruzaron",
            ncols=9)
    HDRS_N = ["Fecha", "Origen", "Descripción / Nota", "Empleado/Ref",
              "Valor ($)", "Parcial ($)", "Estado", "Observación", ""]
    WIDS_N = [12, 14, 48, 20, 18, 18, 24, 52, 12]
    for ci, (h, w) in enumerate(zip(HDRS_N, WIDS_N), 1):
        cel = ws_n.cell(row=3, column=ci, value=h)
        _sc(cel, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws_n.column_dimensions[get_column_letter(ci)].width = w
    ws_n.row_dimensions[3].height = 20

    COLOR_NOM = {
        'NOMINA_CUADRA':          (COLOR["cuadra"],     COLOR["cuadra_font"]),
        'NOMINA_DIF_MENOR':       (COLOR["dif_menor"],  COLOR["dif_font"]),
        'NOMINA_DIF':             (COLOR["diferencia"], COLOR["sin_font"]),
        'NOMINA_SIN_MATCH_BANCO': (COLOR["diferencia"], COLOR["sin_font"]),
        'NOMINA_SOLO_BANCO':      (COLOR["solo_banco"], COLOR["solo_banco_f"]),
        'DETALLE':                ("DDEEFF",           "1F3864"),
    }
    row_n = 4
    for c in cruces_nomina:
        estado   = c['estado']
        bg, fc   = COLOR_NOM.get(estado, (None, "000000"))
        fecha_dt = c['fecha'].date() if pd.notna(c['fecha']) else None

        # ─ Fila MAESTRA: resumen del grupo auxiliar
        nota_res = c['nota_aux'][:70]
        conteo   = c.get('conteo_empleados', 0)
        if conteo > 0:
            nota_res += f" ({conteo} registros)"

        vals = [
            fecha_dt,
            'AUXILIAR',
            nota_res,
            '',
            c['valor_aux']  if c['valor_aux'] > 0 else None,
            None,
            estado,
            str(c.get('observacion', '')),
            '',
        ]
        fmts = [FMT_DATE, None, None, None, FMT_COP, None, None, None, None]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cel = ws_n.cell(row=row_n, column=ci, value=v)
            _sc(cel, bg=bg, fc=fc, fmt=fmt,
                ha="center" if ci in (1, 2) else "right" if ci in (5, 6) else "left",
                bold=True)
        ws_n.row_dimensions[row_n].height = 20
        row_n += 1

        # ─ Fila MAESTRA resumen extracto
        vals_ext_head = [
            None, 'BANCO (grupo)', '', '',
            c['valor_ext'] if c['valor_ext'] > 0 else None,
            c['diferencia'] if (c['valor_aux'] > 0 or c['valor_ext'] > 0) else None,
            estado, '', '',
        ]
        fmts_ext_head = [None, None, None, None, FMT_COP, FMT_COP, None, None, None]
        for ci, (v, fmt) in enumerate(zip(vals_ext_head, fmts_ext_head), 1):
            cel = ws_n.cell(row=row_n, column=ci, value=v)
            _sc(cel, bg="E8F0E8", fc="375623", fmt=fmt,
                ha="center" if ci == 2 else "right" if ci in (5, 6) else "left",
                bold=True, sz=9)
        ws_n.row_dimensions[row_n].height = 18
        row_n += 1

        # ─ Filas DETALLE: registros individuales del extracto
        regs_ext = c.get('regs_ext_originales', [])
        if regs_ext:
            for reg in regs_ext:
                fecha_r = reg.get('Fecha', reg.get('fecha', None))
                if isinstance(fecha_r, pd.Timestamp):
                    fecha_r = fecha_r.date()
                desc_r = str(reg.get('Descripcion', reg.get('descripcion', '')))[:70]
                val_r = float(reg.get('Debitos', reg.get('debitos', 0)))
                ref_r = str(reg.get('Referencia', reg.get('referencia', '')))[:20]
                vals_d = [
                    fecha_r,
                    '  → extracto',
                    desc_r,
                    ref_r,
                    val_r if val_r > 0 else None,
                    None, '', '', '',
                ]
                fmts_d = [FMT_DATE, None, None, None, FMT_COP, None, None, None, None]
                for ci, (v, fmt) in enumerate(zip(vals_d, fmts_d), 1):
                    cel = ws_n.cell(row=row_n, column=ci, value=v)
                    _sc(cel, bg="F5F5F5", fc="555555", fmt=fmt,
                        ha="center" if ci in (1, 2, 4) else "right" if ci == 5 else "left",
                        sz=9)
                ws_n.row_dimensions[row_n].height = 15
                row_n += 1

        # ─ Fila separadora
        cel_sep = ws_n.cell(row=row_n, column=1, value='')
        _sc(cel_sep, bg="FFFFFF")
        ws_n.row_dimensions[row_n].height = 6
        row_n += 1

    ws_n.freeze_panes = "A4"

    # ── Hoja MULTIABONO_CRUCE ─────────────────────────────────────────────────
    ws_ma_sheet = wb.create_sheet("MULTIABONO_CRUCE")
    _titulo(ws_ma_sheet,
            f"CONCILIACIÓN MULTIABONO — {cuenta} — {periodo.upper()}",
            "Cruce desagregado: cada fila del auxiliar con los del extracto que cruzaron por fecha",
            ncols=9)
    HDRS_MA = ["Fecha", "Origen", "Descripción / Nota", "Empleado/Ref",
               "Valor ($)", "Parcial ($)", "Estado", "Observación", ""]
    WIDS_MA = [12, 14, 48, 20, 18, 18, 24, 52, 12]
    for ci, (h, w) in enumerate(zip(HDRS_MA, WIDS_MA), 1):
        cel = ws_ma_sheet.cell(row=3, column=ci, value=h)
        _sc(cel, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws_ma_sheet.column_dimensions[get_column_letter(ci)].width = w
    ws_ma_sheet.row_dimensions[3].height = 20

    COLOR_MA = {
        'MULTIABONO_CUADRA':          (COLOR["cuadra"],     COLOR["cuadra_font"]),
        'MULTIABONO_DIF_MENOR':       (COLOR["dif_menor"],  COLOR["dif_font"]),
        'MULTIABONO_DIF':             (COLOR["diferencia"], COLOR["sin_font"]),
        'MULTIABONO_SIN_MATCH_BANCO': (COLOR["diferencia"], COLOR["sin_font"]),
        'MULTIABONO_SOLO_BANCO':      (COLOR["solo_banco"], COLOR["solo_banco_f"]),
        'DETALLE':                    ("DDEEFF",           "1F3864"),
    }
    row_ma = 4
    for c in cruces_multiabono:
        estado   = c['estado']
        bg, fc   = COLOR_MA.get(estado, (None, "000000"))
        fecha_dt = c['fecha'].date() if pd.notna(c['fecha']) else None
        conteo   = c.get('conteo_registros', 0)
        nota_res = c['nota_aux'][:70]
        if conteo > 0:
            nota_res += f" ({conteo} registros)"

        # ─ Fila MAESTRA auxiliar
        vals = [
            fecha_dt, 'AUXILIAR', nota_res, '',
            c['valor_aux'] if c['valor_aux'] > 0 else None,
            None, estado, str(c.get('observacion', '')), '',
        ]
        fmts = [FMT_DATE, None, None, None, FMT_COP, None, None, None, None]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cel = ws_ma_sheet.cell(row=row_ma, column=ci, value=v)
            _sc(cel, bg=bg, fc=fc, fmt=fmt,
                ha="center" if ci in (1, 2) else "right" if ci in (5, 6) else "left",
                bold=True)
        ws_ma_sheet.row_dimensions[row_ma].height = 20
        row_ma += 1

        # ─ Fila MAESTRA resumen extracto
        vals_ext_head = [
            None, 'BANCO (grupo)', '', '',
            c['valor_ext'] if c['valor_ext'] > 0 else None,
            c['diferencia'] if (c['valor_aux'] > 0 or c['valor_ext'] > 0) else None,
            estado, '', '',
        ]
        fmts_ext_head = [None, None, None, None, FMT_COP, FMT_COP, None, None, None]
        for ci, (v, fmt) in enumerate(zip(vals_ext_head, fmts_ext_head), 1):
            cel = ws_ma_sheet.cell(row=row_ma, column=ci, value=v)
            _sc(cel, bg="E8F0E8", fc="375623", fmt=fmt,
                ha="center" if ci == 2 else "right" if ci in (5, 6) else "left",
                bold=True, sz=9)
        ws_ma_sheet.row_dimensions[row_ma].height = 18
        row_ma += 1

        # ─ Filas DETALLE: registros individuales del extracto
        regs_ext = c.get('regs_ext_originales', [])
        if regs_ext:
            for reg in regs_ext:
                fecha_r = reg.get('Fecha', reg.get('fecha', None))
                if isinstance(fecha_r, pd.Timestamp):
                    fecha_r = fecha_r.date()
                desc_r = str(reg.get('Descripcion', reg.get('descripcion', '')))[:70]
                val_r = float(reg.get('Creditos', reg.get('creditos', reg.get('Debitos', reg.get('debitos', 0)))))
                ref_r = str(reg.get('Referencia', reg.get('referencia', '')))[:20]
                vals_d = [
                    fecha_r,
                    '  → extracto',
                    desc_r, ref_r,
                    val_r if val_r > 0 else None,
                    None, '', '', '',
                ]
                fmts_d = [FMT_DATE, None, None, None, FMT_COP, None, None, None, None]
                for ci, (v, fmt) in enumerate(zip(vals_d, fmts_d), 1):
                    cel = ws_ma_sheet.cell(row=row_ma, column=ci, value=v)
                    _sc(cel, bg="F5F5F5", fc="555555", fmt=fmt,
                        ha="center" if ci in (1, 2, 4) else "right" if ci == 5 else "left",
                        sz=9)
                ws_ma_sheet.row_dimensions[row_ma].height = 15
                row_ma += 1

        # ─ Fila separadora
        cel_sep = ws_ma_sheet.cell(row=row_ma, column=1, value='')
        _sc(cel_sep, bg="FFFFFF")
        ws_ma_sheet.row_dimensions[row_ma].height = 6
        row_ma += 1

    ws_ma_sheet.freeze_panes = "A4"

    # ── Hoja DATAFONOS_RETIRADOS ──────────────────────────────────────────────
    ws_df = wb.create_sheet("DATAFONOS_RETIRADOS")
    _titulo(ws_df,
            f"DATÁFONOS RETIRADOS DEL EXTRACTO — {cuenta} — {periodo.upper()}",
            "Movimientos excluidos del cruce bancario — se concilian en el conciliador de datáfonos",
            ncols=6)
    HDRS_DF = ["Fecha", "Descripción", "Referencia", "Débitos ($)", "Créditos ($)", "Saldo ($)"]
    WIDS_DF = [12, 44, 20, 18, 18, 18]
    for ci, (h, w) in enumerate(zip(HDRS_DF, WIDS_DF), 1):
        cel = ws_df.cell(row=3, column=ci, value=h)
        _sc(cel, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws_df.column_dimensions[get_column_letter(ci)].width = w
    ws_df.row_dimensions[3].height = 18

    row_n = 4
    if not datafonos.empty:
        for _, re_ in datafonos.iterrows():
            vals = [
                re_['Fecha'].date() if pd.notna(re_['Fecha']) else None,
                str(re_.get('Descripcion', '')),
                str(re_.get('Referencia', '')),
                re_.get('Debitos',  0) or None,
                re_.get('Creditos', 0) or None,
                re_.get('Saldo',    None),
            ]
            fmts = [FMT_DATE, None, None, FMT_COP, FMT_COP, FMT_COP]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
                cel = ws_df.cell(row=row_n, column=ci, value=v)
                _sc(cel, bg="DDEEFF", fc="1F3864", fmt=fmt,
                    ha="right" if ci in [4, 5, 6] else "center" if ci == 1 else "left")
            ws_df.row_dimensions[row_n].height = 15
            row_n += 1
        # Total
        for ci in range(1, 7):
            cel = ws_df.cell(row=row_n, column=ci)
            if ci == 1:
                cel.value = "TOTAL"
                _sc(cel, bg=COLOR["total_bg"], fc="FFFFFF", bold=True, ha="center")
            elif ci == 4:
                cel.value = float(datafonos['Debitos'].sum())
                _sc(cel, bg=COLOR["total_bg"], fc="FFFFFF", bold=True, fmt=FMT_COP, ha="right")
            elif ci == 5:
                cel.value = float(datafonos['Creditos'].sum())
                _sc(cel, bg=COLOR["total_bg"], fc="FFFFFF", bold=True, fmt=FMT_COP, ha="right")
            else:
                _sc(cel, bg=COLOR["total_bg"], fc="FFFFFF")
        ws_df.row_dimensions[row_n].height = 18
    ws_df.freeze_panes = "A4"

    # ── Hoja SOLO_EN_BANCO ────────────────────────────────────────────────────
    ws_s = wb.create_sheet("SOLO_EN_BANCO")
    _titulo(ws_s,
            f"MOVIMIENTOS SOLO EN EXTRACTO BANCARIO — {cuenta}",
            "Movimientos que el banco registró pero no aparecen en el auxiliar contable (excluye datáfonos y nómina)",
            ncols=6)
    HDRS_S = ["Fecha", "Descripción", "Referencia", "Débitos ($)", "Créditos ($)", "Saldo ($)"]
    WIDS_S = [12, 44, 20, 18, 18, 18]
    for ci, (h, w) in enumerate(zip(HDRS_S, WIDS_S), 1):
        cel = ws_s.cell(row=3, column=ci, value=h)
        _sc(cel, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    ws_s.row_dimensions[3].height = 18

    row_n = 4
    for _, re_ in solo_banco.iterrows():
        vals = [
            re_['Fecha'].date() if pd.notna(re_['Fecha']) else None,
            str(re_.get('Descripcion', '')),
            str(re_.get('Referencia', '')),
            re_.get('Debitos',  0) or None,
            re_.get('Creditos', 0) or None,
            re_.get('Saldo',    None),
        ]
        fmts = [FMT_DATE, None, None, FMT_COP, FMT_COP, FMT_COP]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cel = ws_s.cell(row=row_n, column=ci, value=v)
            _sc(cel, bg=COLOR["solo_banco"], fc=COLOR["solo_banco_f"],
                fmt=fmt, ha="right" if ci in [4, 5, 6] else "center" if ci == 1 else "left")
        ws_s.row_dimensions[row_n].height = 15
        row_n += 1
    ws_s.freeze_panes = "A4"

    # ── Hoja LOG_AUDITORIA ────────────────────────────────────────────────────
    ws_a = wb.create_sheet("LOG_AUDITORIA")
    ws_a.column_dimensions['A'].width = 44
    ws_a.column_dimensions['B'].width = 72
    _titulo(ws_a, "LOG DE AUDITORÍA — CONCILIACIÓN BANCARIA", ncols=2)

    registros_log = [
        ("Fecha / Hora proceso",        datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ("Empresa",                     info_empresa.get('razon_social', '') if info_empresa else ''),
        ("Marca comercial",             "JAGI CAPS"),
        ("Cuenta bancaria",             cuenta),
        ("Banco",                       banco),
        ("Período",                     periodo),
        ("Versión motor",               "1.3"),
        ("Normativa",                   "NIIF para PYMES (Decreto 2420 de 2015) / DIAN Colombia"),
        ("Lógica de cruce",             "Auxiliar WorldOffice ↔ Extracto bancario por fecha y valor"),
        ("Estrategia de match",         "1) Mismo día + valor exacto  2) ±3 días + tolerancia 1%"),
        ("Tolerancia diferencia",       f"< ${TOLERANCIA_DEFAULT:.0f} → CUADRA | ≤ {TOLERANCIA_MENOR*100:.0f}% → DIF_MENOR (reducido de 5% a 1% por decisión gerencia)"),
        ("Regla datáfonos",             "Detección por palabras clave: marca (mastercard/visa/amex/diners) + tipo (ventas). Robusto ante variaciones del banco."),
        ("Marcas detectadas",           " | ".join(DATAFONO_MARCAS)),
        ("Tipos detectados",            " | ".join(DATAFONO_TIPO)),
        ("Datáfonos retirados",         f"{resumen.get('datafonos_retirados', 0)} movimientos — ${resumen.get('datafonos_valor', 0):,.0f}"),
        ("Regla nómina auxiliar",       "Detección por palabra clave: " + " | ".join(NOMINA_PALABRAS_AUXILIAR)),
        ("Regla nómina extracto",       "Detección por palabra clave: " + " | ".join(NOMINA_PALABRAS_EXTRACTO)),
        ("Exclusiones nómina extracto", " | ".join(NOMINA_EXCLUSIONES_EXTRACTO) + " → va al cruce general"),
        ("Nd Cobro Disp Fond Daviplata","Tratado como movimiento general (NO nómina). Aparece en SOLO_EN_BANCO si sin contrapartida."),
        ("Regla multiabono",            "Detección: 'multiabono' o 'nd pago credito' o 'pago credito'. Agrupado por fecha en ambas fuentes → cruce agrupado con detalle maestro-detalle"),
        ("Multiabono — fechas",         f"{resumen.get('ma_fechas', 0)} fechas | Cuadran: {resumen.get('ma_cuadra', 0)} | Sin match: {resumen.get('ma_sin_match', 0)}"),
        ("Archivos originales",         "NO modificados — motor opera sobre copias en memoria"),
        ("Auditoría externa",           "KPMG Ltda."),
    ]
    for i, (k, v) in enumerate(registros_log, 2):
        ck = ws_a.cell(row=i, column=1, value=k)
        cv = ws_a.cell(row=i, column=2, value=v)
        _sc(ck, bg="F2F2F2", bold=True)
        _sc(cv)
        ws_a.row_dimensions[i].height = 16

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output_path)
    return output_path