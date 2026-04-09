"""
Automatización de Conciliación Bancaria - JWG
Cruza el Libro Auxiliar con el Extracto Bancario siguiendo las reglas contables definidas.

VERSION: 2026-04-07 (validación mejorada)
- ✅ Datáfono: desfase hasta 5 días hábiles (fines de semana/festivos)
- ✅ RC: validación de tercero PENDIENTE (reunión con contabilidad pendiente)
- ✅ Logging: campo 'Regla Aplicada' en cada conciliación (auditoría)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import re
import threading

# ──────────────────────────────────────────────
# FESTIVOS COLOMBIANOS (años 2024-2026)
# ──────────────────────────────────────────────
FESTIVOS_CO = {
    date(2025, 1, 1), date(2025, 1, 6), date(2025, 3, 24), date(2025, 4, 17),
    date(2025, 4, 18), date(2025, 5, 1), date(2025, 6, 2), date(2025, 6, 23),
    date(2025, 6, 30), date(2025, 8, 7), date(2025, 8, 18), date(2025, 10, 13),
    date(2025, 11, 3), date(2025, 11, 17), date(2025, 12, 8), date(2025, 12, 25),
    date(2026, 1, 1), date(2026, 1, 5), date(2026, 3, 23), date(2026, 4, 2),
    date(2026, 4, 3), date(2026, 5, 1), date(2026, 6, 1), date(2026, 6, 22),
    date(2026, 6, 29), date(2026, 8, 7), date(2026, 8, 17), date(2026, 10, 12),
    date(2026, 11, 2), date(2026, 11, 16), date(2026, 12, 8), date(2026, 12, 25),
}

# ──────────────────────────────────────────────
# CONSTANTES DE CLASIFICACIÓN
# ──────────────────────────────────────────────
GASTOS_BANCARIOS_KEYWORDS = [
    "ABONO INTERESES AHORROS", "AJUSTE INTERES AHORROS DB",
    "COMIS CONSIGNACION CB", "COMISION TRASLADO OTROS BANCOS",
    "IMPTO GOBIERNO 4X1000", "IVA COMIS TRASLADO OTROS BCOS",
    "REV IMPTO GOBIERNO 4X1000", "VALOR IVA",
]

MESES_EXCEPCION = {1, 2, 3}  # enero, febrero, marzo

# ──────────────────────────────────────────────
# LÓGICA DE CONCILIACIÓN
# ──────────────────────────────────────────────

def normalizar_fecha(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def clasificar_doc(doc_num, nota=""):
    """
    Retorna el tipo de documento según Doc Num + Nota del auxiliar.
    La Nota es la fuente principal para NC (notas de contabilidad).
    """
    if not doc_num:
        return "OTRO"
    d   = str(doc_num).upper()
    n   = str(nota).upper() if nota else ""

    if "CE W" in d or "CE AJ" in d:
        return "CE"

    if "NC" in d:
        # Clasificar por contenido de la NOTA (fuente primaria)
        if "EFECTIVO" in n:
            return "NC_EFECTIVO"
        if "QR" in n:
            return "NC_QR"
        if "TRANSFERENCIA" in n or "TRANSFER" in n:
            return "NC_TRANSFERENCIA"
        if "DATAFONO" in n or "DATÁFONO" in n or "REDEBAN" in n:
            return "NC_DATAFONO"
        if any(k in n for k in [
            "GASTO", "BANCARIO", "COMIS", "4X1000", "VALOR IVA", "IVA COMIS", "INTERESES"
        ]):
            return "NC_GASTOS_BANCARIOS"
        if "TDC" in n or "TARJETA" in n:
            return "NC_TRANSFERENCIA"   # pagos de tarjeta = transferencia
        if "DEVOLUCION" in n:
            return "NC_EFECTIVO"
        return "NC_OTRO"

    if "RC" in d:
        return "RC"
    return "OTRO"


def es_gasto_bancario_extracto(desc):
    desc_up = str(desc).upper()
    return any(k in desc_up for k in GASTOS_BANCARIOS_KEYWORDS)


def es_qr_extracto(desc):
    desc_up = str(desc).upper()
    return "PAGO QR" in desc_up or "TRANSF QR" in desc_up


def es_consignacion_extracto(desc):
    desc_up = str(desc).upper()
    return any(x in desc_up for x in [
        "CONSIGNACION CORRESPONSAL CB", "CONSIG NACIONAL", "CONSIG LOCAL",
        "CONSIG EFECTIVO", "CONSIGNACION EFECTIVO"
    ])


def es_transferencia_sucursal(desc):
    desc_up = str(desc).upper()
    return "TRANSFERENCIA CTA SUC" in desc_up or "TRANSFERENCIAS CTA SUC" in desc_up


def es_recaudo(desc):
    desc_up = str(desc).upper()
    return "PAGO INTERBANC" in desc_up or "PAGO DE PROV" in desc_up


def validar_tercero_rc(tercero_aux, descripcion_ext):
    """
    TODO: Definir con contabilidad cómo extraer/validar el tercero en extracto.
    Por ahora, retorna True (acepta todo) para no bloquear el proceso.
    Se debe implementar después de la reunión.
    """
    # Posibles estrategias (a definir):
    # 1. Buscar NIT en descripción del extracto
    # 2. Comparar con lista de clientes/proveedores
    # 3. Usar campo TERCERO si existe en el extracto
    return True


def dias_habiles_entre(d1, d2):
    """Días hábiles (lunes-sábado, excluyendo festivos CO) entre dos fechas."""
    if d1 > d2:
        d1, d2 = d2, d1
    count = 0
    cur = d1
    while cur <= d2:
        if cur.weekday() < 6 and cur not in FESTIVOS_CO:  # 0-5 = lun-sab
            count += 1
        cur = date.fromordinal(cur.toordinal() + 1)
    return count


def es_desfase_datafono_aceptable(fecha_aux, fecha_ext, max_dias_habiles=5):
    """
    NC_DATAFONO permite desfase por fines de semana/festivos.
    Requiere: mismo mes + diferencia <= max_dias_hábiles.
    """
    if not mismo_mes(fecha_aux, fecha_ext):
        return False
    return dias_habiles_entre(fecha_aux, fecha_ext) <= max_dias_habiles


def mismo_mes(d1, d2):
    return d1 is not None and d2 is not None and d1.year == d2.year and d1.month == d2.month


def cargar_auxiliar(path):
    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date
    df["Debitos"] = pd.to_numeric(df["Debitos"], errors="coerce").fillna(0)
    df["Creditos"] = pd.to_numeric(df["Creditos"], errors="coerce").fillna(0)
    df["_tipo_doc"] = df.apply(lambda r: clasificar_doc(r["Doc Num"], r.get("Nota", "")), axis=1)
    df["_idx"] = range(len(df))
    # Monto neto desde perspectiva banco: débito = entrada (+), crédito = salida (-)
    df["_monto_banco"] = df["Debitos"] - df["Creditos"]
    return df


def cargar_extracto(path):
    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce").dt.date
    df["VALOR"] = pd.to_numeric(df["VALOR"], errors="coerce").fillna(0)
    df["_idx"] = range(len(df))
    return df


def conciliar(df_aux, df_ext, mes_filtro=None):
    """
    Ejecuta la conciliación completa según reglas de CLAUDE.md.

    Pasos implementados:
    1. CE W (Comprobante de egreso) → valor exacto
    2. NC_QR (Ingresos QR) → valor + fecha de NOTA
    3. NC_EFECTIVO (Consignación corresponsal) → valor + fecha mes
    4. NC_TRANSFERENCIA → valor + fecha mes
    5. NC_DATAFONO → valor + desfase hasta 5 días hábiles
    6. NC_GASTOS_BANCARIOS → valor exacto (lista cerrada)
    7. RC (Recaudo cartera) → valor + fecha mes (tercero pendiente validar)
    8. Excepción ene-mar → valor + fecha idéntica

    mes_filtro: int (1-12) o None para todos los meses.
    Retorna: (conciliadas, pend_aux, pend_banco, errores)
    """
    # Filtrar por mes si se especifica
    if mes_filtro:
        df_aux = df_aux[df_aux["Fecha"].apply(lambda d: d is not None and d.month == mes_filtro)].copy()
        df_ext = df_ext[df_ext["FECHA"].apply(lambda d: d is not None and d.month == mes_filtro)].copy()

    usados_aux = set()
    usados_ext = set()

    conciliadas = []
    errores_clasificacion = []

    # ── PASO 1: CE W → cruza por valor exacto ──
    ce_aux = df_aux[df_aux["_tipo_doc"] == "CE"].copy()
    for _, row_a in ce_aux.iterrows():
        valor_bus = abs(row_a["_monto_banco"])
        fecha_a = row_a["Fecha"]
        nota_a = str(row_a.get("Nota", "")).upper()

        # Excepción: LIQUIDACION DEFINITIVA → CE por valor exacto
        if "LIQUIDACION DEFINITIVA" in nota_a:
            candidatos = df_ext[
                (~df_ext["_idx"].isin(usados_ext)) &
                (df_ext["VALOR"].abs() == valor_bus)
            ]
        else:
            candidatos = df_ext[
                (~df_ext["_idx"].isin(usados_ext)) &
                (df_ext["VALOR"] < 0) &
                (df_ext["VALOR"].abs() == valor_bus)
            ]

        # Aplicar desfase SUFI
        def es_sufi(desc):
            return "DEBITO OBLIGACION SUFI" in str(desc).upper()

        candidatos_mes = candidatos[candidatos["FECHA"].apply(lambda d: mismo_mes(d, fecha_a))]
        if candidatos_mes.empty:
            candidatos_mes = candidatos

        for _, row_e in candidatos_mes.iterrows():
            if row_e["_idx"] in usados_ext:
                continue
            usados_aux.add(row_a["_idx"])
            usados_ext.add(row_e["_idx"])
            conciliadas.append({
                "Tipo": "CE",
                "Fecha Aux": fecha_a,
                "Doc Num": row_a.get("Doc Num"),
                "Tercero": row_a.get("Tercero"),
                "Nota Aux": row_a.get("Nota"),
                "Valor Aux": row_a["_monto_banco"],
                "Fecha Banco": row_e["FECHA"],
                "Descripción Banco": row_e["DESCRIPCIÓN"],
                "Valor Banco": row_e["VALOR"],
                "Regla Aplicada": "Valor exacto" if "LIQUIDACION DEFINITIVA" in nota_a else "Valor exacto (CE)",
            })

    # ── PASO 2: NC_QR → fecha de NOTA, valor exacto ──
    qr_aux = df_aux[(df_aux["_tipo_doc"] == "NC_QR") & (~df_aux["_idx"].isin(usados_aux))].copy()
    qr_ext = df_ext[(~df_ext["_idx"].isin(usados_ext)) & (df_ext["VALOR"] > 0)].copy()
    qr_ext_fil = qr_ext[qr_ext["DESCRIPCIÓN"].apply(es_qr_extracto)]

    for _, row_a in qr_aux.iterrows():
        valor_bus = abs(row_a["_monto_banco"])
        nota_a = str(row_a.get("Nota", ""))
        # Extraer fecha de nota si existe (formato DD/MM/YYYY o YYYY-MM-DD)
        fecha_nota = row_a["Fecha"]  # fallback
        match = re.search(r'(\d{2})[/\-](\d{2})[/\-](\d{2,4})', nota_a)
        if match:
            try:
                g = match.groups()
                yr = int(g[2]) if len(g[2]) == 4 else 2000 + int(g[2])
                fecha_nota = date(yr, int(g[1]), int(g[0]))
            except Exception:
                pass

        candidatos = qr_ext_fil[
            (~qr_ext_fil["_idx"].isin(usados_ext)) &
            (qr_ext_fil["VALOR"].abs() == valor_bus) &
            (qr_ext_fil["FECHA"].apply(lambda d: mismo_mes(d, fecha_nota)))
        ]

        for _, row_e in candidatos.iterrows():
            if row_e["_idx"] in usados_ext:
                continue
            usados_aux.add(row_a["_idx"])
            usados_ext.add(row_e["_idx"])
            conciliadas.append({
                "Tipo": "NC_QR",
                "Fecha Aux": row_a["Fecha"],
                "Doc Num": row_a.get("Doc Num"),
                "Tercero": row_a.get("Tercero"),
                "Nota Aux": row_a.get("Nota"),
                "Valor Aux": row_a["_monto_banco"],
                "Fecha Banco": row_e["FECHA"],
                "Descripción Banco": row_e["DESCRIPCIÓN"],
                "Valor Banco": row_e["VALOR"],
                "Regla Aplicada": "Valor exacto + fecha NOTA",
            })
            break

    # ── PASO 3: NC_EFECTIVO → consignaciones ──
    ef_aux = df_aux[(df_aux["_tipo_doc"] == "NC_EFECTIVO") & (~df_aux["_idx"].isin(usados_aux))].copy()
    ef_ext = df_ext[
        (~df_ext["_idx"].isin(usados_ext)) &
        (df_ext["VALOR"] > 0) &
        (df_ext["DESCRIPCIÓN"].apply(es_consignacion_extracto))
    ].copy()

    for _, row_a in ef_aux.iterrows():
        valor_bus = abs(row_a["_monto_banco"])
        fecha_a = row_a["Fecha"]
        candidatos = ef_ext[
            (~ef_ext["_idx"].isin(usados_ext)) &
            (ef_ext["VALOR"].abs() == valor_bus) &
            (ef_ext["FECHA"].apply(lambda d: mismo_mes(d, fecha_a)))
        ]
        for _, row_e in candidatos.iterrows():
            if row_e["_idx"] in usados_ext:
                continue
            usados_aux.add(row_a["_idx"])
            usados_ext.add(row_e["_idx"])
            conciliadas.append({
                "Tipo": "NC_EFECTIVO",
                "Fecha Aux": fecha_a,
                "Doc Num": row_a.get("Doc Num"),
                "Tercero": row_a.get("Tercero"),
                "Nota Aux": row_a.get("Nota"),
                "Valor Aux": row_a["_monto_banco"],
                "Fecha Banco": row_e["FECHA"],
                "Descripción Banco": row_e["DESCRIPCIÓN"],
                "Valor Banco": row_e["VALOR"],
                "Regla Aplicada": "Valor exacto (consignación)",
            })
            break

    # ── PASO 4: NC_TRANSFERENCIA ──
    tr_aux = df_aux[(df_aux["_tipo_doc"] == "NC_TRANSFERENCIA") & (~df_aux["_idx"].isin(usados_aux))].copy()
    tr_ext = df_ext[
        (~df_ext["_idx"].isin(usados_ext)) &
        (df_ext["VALOR"] > 0) &
        (df_ext["DESCRIPCIÓN"].apply(es_transferencia_sucursal))
    ].copy()

    for _, row_a in tr_aux.iterrows():
        valor_bus = abs(row_a["_monto_banco"])
        fecha_a = row_a["Fecha"]
        candidatos = tr_ext[
            (~tr_ext["_idx"].isin(usados_ext)) &
            (tr_ext["VALOR"].abs() == valor_bus) &
            (tr_ext["FECHA"].apply(lambda d: mismo_mes(d, fecha_a)))
        ]
        for _, row_e in candidatos.iterrows():
            if row_e["_idx"] in usados_ext:
                continue
            usados_aux.add(row_a["_idx"])
            usados_ext.add(row_e["_idx"])
            conciliadas.append({
                "Tipo": "NC_TRANSFERENCIA",
                "Fecha Aux": fecha_a,
                "Doc Num": row_a.get("Doc Num"),
                "Tercero": row_a.get("Tercero"),
                "Nota Aux": row_a.get("Nota"),
                "Valor Aux": row_a["_monto_banco"],
                "Fecha Banco": row_e["FECHA"],
                "Descripción Banco": row_e["DESCRIPCIÓN"],
                "Valor Banco": row_e["VALOR"],
                "Regla Aplicada": "Valor exacto (transferencia)",
            })
            break

    # ── PASO 5: NC_DATAFONO → desfase hasta 5 días hábiles ──
    df_aux2 = df_aux[(df_aux["_tipo_doc"] == "NC_DATAFONO") & (~df_aux["_idx"].isin(usados_aux))].copy()
    df_ext2 = df_ext[
        (~df_ext["_idx"].isin(usados_ext)) &
        (df_ext["VALOR"] > 0) &
        (df_ext["DESCRIPCIÓN"].str.upper().str.contains("ABONO NETO REDEBAN", na=False))
    ].copy()

    for _, row_a in df_aux2.iterrows():
        valor_bus = abs(row_a["_monto_banco"])
        fecha_a = row_a["Fecha"]
        candidatos = df_ext2[
            (~df_ext2["_idx"].isin(usados_ext)) &
            (df_ext2["VALOR"].abs() == valor_bus)
        ]
        # Validar desfase de días hábiles (mismo mes + hasta 5 días hábiles)
        for _, row_e in candidatos.iterrows():
            if row_e["_idx"] in usados_ext:
                continue
            if not es_desfase_datafono_aceptable(fecha_a, row_e["FECHA"], max_dias_habiles=5):
                continue
            usados_aux.add(row_a["_idx"])
            usados_ext.add(row_e["_idx"])
            conciliadas.append({
                "Tipo": "NC_DATAFONO",
                "Fecha Aux": fecha_a,
                "Doc Num": row_a.get("Doc Num"),
                "Tercero": row_a.get("Tercero"),
                "Nota Aux": row_a.get("Nota"),
                "Valor Aux": row_a["_monto_banco"],
                "Fecha Banco": row_e["FECHA"],
                "Descripción Banco": row_e["DESCRIPCIÓN"],
                "Valor Banco": row_e["VALOR"],
                "Regla Aplicada": "Desfase días hábiles (max 5)",
            })
            break

    # ── PASO 6: NC_GASTOS_BANCARIOS → agrupados mensualmente ──
    gb_aux = df_aux[(df_aux["_tipo_doc"] == "NC_GASTOS_BANCARIOS") & (~df_aux["_idx"].isin(usados_aux))].copy()
    gb_ext = df_ext[
        (~df_ext["_idx"].isin(usados_ext)) &
        (df_ext["DESCRIPCIÓN"].apply(es_gasto_bancario_extracto))
    ].copy()

    for _, row_a in gb_aux.iterrows():
        valor_bus = abs(row_a["_monto_banco"])
        fecha_a = row_a["Fecha"]
        candidatos = gb_ext[
            (~gb_ext["_idx"].isin(usados_ext)) &
            (gb_ext["VALOR"].abs() == valor_bus) &
            (gb_ext["FECHA"].apply(lambda d: mismo_mes(d, fecha_a)))
        ]
        for _, row_e in candidatos.iterrows():
            if row_e["_idx"] in usados_ext:
                continue
            usados_aux.add(row_a["_idx"])
            usados_ext.add(row_e["_idx"])
            conciliadas.append({
                "Tipo": "NC_GASTOS_BANCARIOS",
                "Fecha Aux": fecha_a,
                "Doc Num": row_a.get("Doc Num"),
                "Tercero": row_a.get("Tercero"),
                "Nota Aux": row_a.get("Nota"),
                "Valor Aux": row_a["_monto_banco"],
                "Fecha Banco": row_e["FECHA"],
                "Descripción Banco": row_e["DESCRIPCIÓN"],
                "Valor Banco": row_e["VALOR"],
                "Regla Aplicada": "Valor exacto (gasto bancario)",
            })
            break

    # ── PASO 7: RC → recaudo cartera (con validación de tercero pendiente) ──
    rc_aux = df_aux[(df_aux["_tipo_doc"] == "RC") & (~df_aux["_idx"].isin(usados_aux))].copy()
    rc_ext = df_ext[
        (~df_ext["_idx"].isin(usados_ext)) &
        (df_ext["VALOR"] > 0) &
        (df_ext["DESCRIPCIÓN"].apply(es_recaudo))
    ].copy()

    for _, row_a in rc_aux.iterrows():
        valor_bus = abs(row_a["_monto_banco"])
        fecha_a = row_a["Fecha"]
        tercero_a = str(row_a.get("Tercero", "")).upper().strip()
        candidatos = rc_ext[
            (~rc_ext["_idx"].isin(usados_ext)) &
            (rc_ext["VALOR"].abs() == valor_bus) &
            (rc_ext["FECHA"].apply(lambda d: mismo_mes(d, fecha_a)))
        ]
        for _, row_e in candidatos.iterrows():
            if row_e["_idx"] in usados_ext:
                continue
            # TODO: Activar validación de tercero después de reunión con contabilidad
            # if not validar_tercero_rc(tercero_a, row_e["DESCRIPCIÓN"]):
            #     continue
            usados_aux.add(row_a["_idx"])
            usados_ext.add(row_e["_idx"])
            conciliadas.append({
                "Tipo": "RC",
                "Fecha Aux": fecha_a,
                "Doc Num": row_a.get("Doc Num"),
                "Tercero": row_a.get("Tercero"),
                "Nota Aux": row_a.get("Nota"),
                "Valor Aux": row_a["_monto_banco"],
                "Fecha Banco": row_e["FECHA"],
                "Descripción Banco": row_e["DESCRIPCIÓN"],
                "Valor Banco": row_e["VALOR"],
                "Regla Aplicada": "Valor exacto (recaudo cartera)" + (" [tercero validado]" if False else " [sin validar tercero]"),
            })
            break

    # ── PASO 8: Excepción enero-feb-mar (QR / Transferencias / Nequi / Virtual Pyme) ──
    if mes_filtro and mes_filtro in MESES_EXCEPCION:
        tipos_excepcion = {"NC_QR", "NC_TRANSFERENCIA", "NC_OTRO"}
        exc_aux = df_aux[
            (df_aux["_tipo_doc"].isin(tipos_excepcion)) &
            (~df_aux["_idx"].isin(usados_aux))
        ].copy()
        exc_ext = df_ext[
            (~df_ext["_idx"].isin(usados_ext)) &
            (df_ext["VALOR"] > 0)
        ].copy()

        for _, row_a in exc_aux.iterrows():
            valor_bus = abs(row_a["_monto_banco"])
            fecha_a = row_a["Fecha"]
            candidatos = exc_ext[
                (~exc_ext["_idx"].isin(usados_ext)) &
                (exc_ext["VALOR"].abs() == valor_bus) &
                (exc_ext["FECHA"] == fecha_a)
            ]
            for _, row_e in candidatos.iterrows():
                if row_e["_idx"] in usados_ext:
                    continue
                usados_aux.add(row_a["_idx"])
                usados_ext.add(row_e["_idx"])
                conciliadas.append({
                    "Tipo": "EXCEPCION_CRUCE",
                    "Fecha Aux": fecha_a,
                    "Doc Num": row_a.get("Doc Num"),
                    "Tercero": row_a.get("Tercero"),
                    "Nota Aux": row_a.get("Nota"),
                    "Valor Aux": row_a["_monto_banco"],
                    "Fecha Banco": row_e["FECHA"],
                    "Descripción Banco": row_e["DESCRIPCIÓN"],
                    "Valor Banco": row_e["VALOR"],
                    "Regla Aplicada": "Excepción ene-mar (valor y fecha soporte)",
                })
                break

    # ── Detectar errores de clasificación ──
    # Regla: QR en extracto no debe ser gastos bancarios
    for _, row_e in df_ext.iterrows():
        desc = str(row_e.get("DESCRIPCIÓN", "")).upper()
        if "PAGO QR" in desc and es_gasto_bancario_extracto(desc):
            errores_clasificacion.append({
                "Tipo Error": "QR clasificado como gasto bancario",
                "Fecha Banco": row_e["FECHA"],
                "Descripción": row_e["DESCRIPCIÓN"],
                "Valor": row_e["VALOR"],
                "Explicación": "Los ingresos QR nunca son gastos bancarios (Regla 3)",
            })

    # Pendientes auxiliar
    pend_aux = df_aux[~df_aux["_idx"].isin(usados_aux)].copy()
    pend_aux_out = pend_aux[[
        "Fecha", "Tercero", "Nota", "Doc Num", "Debitos", "Creditos", "_monto_banco", "_tipo_doc"
    ]].rename(columns={"_monto_banco": "Monto Banco", "_tipo_doc": "Tipo Doc"})

    # Pendientes extracto
    pend_ext = df_ext[~df_ext["_idx"].isin(usados_ext)].copy()
    pend_ext_out = pend_ext[["FECHA", "DESCRIPCIÓN", "VALOR"]].rename(columns={
        "FECHA": "Fecha", "DESCRIPCIÓN": "Descripción", "VALOR": "Valor"
    })

    return (
        pd.DataFrame(conciliadas),
        pend_aux_out.reset_index(drop=True),
        pend_ext_out.reset_index(drop=True),
        pd.DataFrame(errores_clasificacion),
    )


# ──────────────────────────────────────────────
# EXPORTAR A EXCEL CON FORMATO
# ──────────────────────────────────────────────

def exportar_excel(df_conc, df_paux, df_pext, df_err, path_out):
    wb = openpyxl.Workbook()

    COLOR_VERDE  = "C6EFCE"
    COLOR_AMARILLO = "FFEB9C"
    COLOR_AZUL   = "BDD7EE"
    COLOR_ROJO   = "FFC7CE"
    COLOR_HEADER = "2E4057"
    FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11)
    ALIN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def escribir_hoja(ws, df, titulo, color_fill):
        fill_header = PatternFill("solid", fgColor=COLOR_HEADER)
        fill_data   = PatternFill("solid", fgColor=color_fill)

        ws.append([titulo])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))
        cell = ws.cell(1, 1)
        cell.font = Font(bold=True, color="FFFFFF", size=13)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = ALIN_CENTER
        ws.row_dimensions[1].height = 25

        ws.append(list(df.columns))
        for col_num, col_name in enumerate(df.columns, 1):
            c = ws.cell(2, col_num)
            c.font = FONT_HEADER
            c.fill = fill_header
            c.alignment = ALIN_CENTER
            c.border = border

        for r_idx, row in enumerate(df.itertuples(index=False), 3):
            for c_idx, val in enumerate(row, 1):
                c = ws.cell(r_idx, c_idx)
                # Formatear fechas y montos
                if isinstance(val, date):
                    c.value = val.strftime("%d/%m/%Y") if pd.notna(val) else ""
                elif isinstance(val, float) and not pd.isna(val):
                    c.value = val
                    c.number_format = '#,##0.00'
                else:
                    c.value = val if not (isinstance(val, float) and pd.isna(val)) else ""
                c.fill = fill_data
                c.border = border
                c.alignment = Alignment(vertical="center")

        # Autoajustar columnas
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    # Hoja Resumen
    ws0 = wb.active
    ws0.title = "📊 Resumen"
    ws0.sheet_view.showGridLines = False
    ws0["A1"] = "CONCILIACIÓN BANCARIA"
    ws0["A1"].font = Font(bold=True, size=16, color="2E4057")
    ws0["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws0["A2"].font = Font(italic=True, color="888888")

    resumen = [
        ("✅ Partidas conciliadas", len(df_conc), COLOR_VERDE),
        ("🟡 Pendientes en Auxiliar", len(df_paux), COLOR_AMARILLO),
        ("🔵 Pendientes en Banco", len(df_pext), COLOR_AZUL),
        ("🔴 Errores detectados", len(df_err), COLOR_ROJO),
    ]
    for i, (label, count, color) in enumerate(resumen, 4):
        ws0[f"A{i}"] = label
        ws0[f"B{i}"] = count
        ws0[f"A{i}"].font = Font(bold=True, size=12)
        ws0[f"B{i}"].font = Font(bold=True, size=12)
        ws0[f"A{i}"].fill = PatternFill("solid", fgColor=color)
        ws0[f"B{i}"].fill = PatternFill("solid", fgColor=color)
        ws0[f"A{i}"].border = border
        ws0[f"B{i}"].border = border
    ws0.column_dimensions["A"].width = 35
    ws0.column_dimensions["B"].width = 12

    if df_conc.empty:
        df_conc = pd.DataFrame(columns=[
            "Tipo","Fecha Aux","Doc Num","Tercero","Nota Aux",
            "Valor Aux","Fecha Banco","Descripción Banco","Valor Banco",
            "Regla Aplicada"
        ])
    if df_paux.empty:
        df_paux = pd.DataFrame(columns=["Fecha","Tercero","Nota","Doc Num","Debitos","Creditos","Monto Banco","Tipo Doc"])
    if df_pext.empty:
        df_pext = pd.DataFrame(columns=["Fecha","Descripción","Valor"])
    if df_err.empty:
        df_err = pd.DataFrame(columns=["Tipo Error","Fecha Banco","Descripción","Valor","Explicación"])

    ws1 = wb.create_sheet("✅ Conciliadas")
    escribir_hoja(ws1, df_conc, "✅ PARTIDAS CONCILIADAS", COLOR_VERDE)

    ws2 = wb.create_sheet("🟡 Pendientes Auxiliar")
    escribir_hoja(ws2, df_paux, "🟡 PENDIENTES EN AUXILIAR", COLOR_AMARILLO)

    ws3 = wb.create_sheet("🔵 Pendientes Banco")
    escribir_hoja(ws3, df_pext, "🔵 PENDIENTES EN BANCO", COLOR_AZUL)

    ws4 = wb.create_sheet("🔴 Errores")
    escribir_hoja(ws4, df_err, "🔴 ERRORES DE CLASIFICACIÓN CONTABLE", COLOR_ROJO)

    wb.save(path_out)


# ──────────────────────────────────────────────
# INTERFAZ GRÁFICA
# ──────────────────────────────────────────────

class ConciliacionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Conciliación Bancaria Automatizada")
        self.root.geometry("700x580")
        self.root.resizable(False, False)
        self.root.configure(bg="#F0F4F8")

        self.path_aux = tk.StringVar()
        self.path_ext = tk.StringVar()
        self.path_out = tk.StringVar()
        self.mes_var  = tk.StringVar(value="Todos los meses")

        self._build_ui()

    def _build_ui(self):
        # Título
        header = tk.Frame(self.root, bg="#2E4057", height=65)
        header.pack(fill="x")
        tk.Label(
            header, text="🏦  Conciliación Bancaria Automatizada",
            font=("Segoe UI", 16, "bold"), bg="#2E4057", fg="white"
        ).pack(pady=14)

        # Cuerpo
        body = tk.Frame(self.root, bg="#F0F4F8", padx=30, pady=20)
        body.pack(fill="both", expand=True)

        self._file_row(body, "📂  Libro Auxiliar (.xlsx):", self.path_aux, 0)
        self._file_row(body, "📂  Extracto Bancario (.xlsx):", self.path_ext, 1)
        self._file_row(body, "💾  Guardar resultado en:", self.path_out, 2, save=True)

        # Selector de mes
        tk.Label(body, text="📅  Mes a conciliar:", font=("Segoe UI", 11),
                 bg="#F0F4F8", fg="#333").grid(row=3, column=0, sticky="w", pady=(14, 4))
        meses = ["Todos los meses", "Enero", "Febrero", "Marzo", "Abril", "Mayo",
                 "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        cb = ttk.Combobox(body, textvariable=self.mes_var, values=meses,
                          state="readonly", font=("Segoe UI", 11), width=22)
        cb.grid(row=3, column=1, sticky="w", pady=(14, 4))

        # Botón ejecutar
        btn = tk.Button(
            body, text="▶  Ejecutar Conciliación",
            font=("Segoe UI", 13, "bold"),
            bg="#1DB954", fg="white", activebackground="#17a349",
            relief="flat", cursor="hand2", padx=20, pady=10,
            command=self._ejecutar
        )
        btn.grid(row=4, column=0, columnspan=2, pady=20)

        # Log
        tk.Label(body, text="Registro de proceso:", font=("Segoe UI", 10, "bold"),
                 bg="#F0F4F8", fg="#555").grid(row=5, column=0, columnspan=2, sticky="w")
        self.log_box = tk.Text(body, height=9, font=("Consolas", 10),
                               bg="#1E1E2E", fg="#CDD6F4", state="disabled",
                               relief="flat", bd=0)
        self.log_box.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(4, 0))
        body.columnconfigure(1, weight=1)

        # Barra de progreso
        self.progress = ttk.Progressbar(body, mode="indeterminate")
        self.progress.grid(row=7, column=0, columnspan=2, sticky="ew", pady=(8, 0))

    def _file_row(self, parent, label, var, row, save=False):
        tk.Label(parent, text=label, font=("Segoe UI", 11),
                 bg="#F0F4F8", fg="#333").grid(row=row, column=0, sticky="w", pady=6)
        frame = tk.Frame(parent, bg="#F0F4F8")
        frame.grid(row=row, column=1, sticky="ew", pady=6)
        tk.Entry(frame, textvariable=var, font=("Segoe UI", 10), width=32,
                 relief="flat", bg="white", bd=1).pack(side="left", padx=(0, 6))
        cmd = (lambda v=var: self._save_dialog(v)) if save else (lambda v=var: self._open_dialog(v))
        tk.Button(frame, text="📁", font=("Segoe UI", 11),
                  bg="#E0E8F0", relief="flat", cursor="hand2",
                  command=cmd).pack(side="left")

    def _open_dialog(self, var):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            var.set(path)

    def _save_dialog(self, var):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Conciliacion_Bancaria.xlsx"
        )
        if path:
            var.set(path)

    def _log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _ejecutar(self):
        if not self.path_aux.get() or not self.path_ext.get() or not self.path_out.get():
            messagebox.showerror("Campos faltantes", "Por favor completa todos los campos.")
            return
        self.progress.start(10)
        threading.Thread(target=self._run_proceso, daemon=True).start()

    def _run_proceso(self):
        try:
            self._log("Cargando Auxiliar...")
            df_aux = cargar_auxiliar(self.path_aux.get())
            self._log(f"  → {len(df_aux)} registros en Auxiliar")

            self._log("Cargando Extracto Bancario...")
            df_ext = cargar_extracto(self.path_ext.get())
            self._log(f"  → {len(df_ext)} movimientos en Extracto")

            meses_map = {
                "Todos los meses": None,
                "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
                "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8,
                "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12,
            }
            mes_filtro = meses_map.get(self.mes_var.get())

            self._log("Ejecutando conciliación...")
            df_conc, df_paux, df_pext, df_err = conciliar(df_aux, df_ext, mes_filtro)

            self._log(f"✅ Conciliadas:           {len(df_conc)}")
            self._log(f"🟡 Pendientes Auxiliar:   {len(df_paux)}")
            self._log(f"🔵 Pendientes Banco:      {len(df_pext)}")
            self._log(f"🔴 Errores detectados:    {len(df_err)}")

            self._log("Exportando resultado a Excel...")
            exportar_excel(df_conc, df_paux, df_pext, df_err, self.path_out.get())
            self._log(f"✔ Archivo guardado: {self.path_out.get()}")

            self.root.after(0, lambda: messagebox.showinfo(
                "¡Proceso completado!",
                f"Conciliación finalizada.\n\n"
                f"✅ Conciliadas:         {len(df_conc)}\n"
                f"🟡 Pendientes Auxiliar: {len(df_paux)}\n"
                f"🔵 Pendientes Banco:    {len(df_pext)}\n"
                f"🔴 Errores:             {len(df_err)}\n\n"
                f"Resultado guardado en:\n{self.path_out.get()}"
            ))
        except Exception as e:
            self._log(f"❌ ERROR: {e}")
            self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.root.after(0, self.progress.stop)


def main():
    root = tk.Tk()
    app = ConciliacionApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
