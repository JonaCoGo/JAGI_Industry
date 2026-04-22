"""
Motor de Conciliación de Datafonos — v1.5
Empresa: GIRALDO GIRALDO JAIME WILSON  |  Normativa: NIIF PYMES / DIAN Colombia

NOVEDADES v1.5:
  • Soporte para formato antiguo del auxiliar (fecha en Nota + sede en columna Tercero)
  • Detección automática de formato por fila — un mismo auxiliar puede mezclar ambos
  • Extracción flexible del día desde la Nota (fecha completa, parcial, o número al inicio)
  • Limpieza de prefijos en Tercero: quita CLIENTES VENTAS / CLIENTES VENTA / CLIENTES
    (PLAZA, PARQUE, VIVA, LOCAL, etc. se conservan — son parte del nombre de la sede)
  • Lista oficial de 25 sedes para normalizar nombres entre formatos y años
  • Hoja PENDIENTES en el Excel: muestra lo que no cuadró en auxiliar y en datafono
  • Nunca usa la columna Fecha del auxiliar para extraer el día
"""

import pandas as pd
import re
import os
import unicodedata
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ──────────────────────────────────────────────────────────────────
COLOR = {
    "header_bg":    "1F3864", "header_font":  "FFFFFF",
    "cuadra":       "C6EFCE", "cuadra_font":  "276221",
    "dif_menor":    "FFEB9C", "dif_font":     "9C5700",
    "diferencia":   "FFC7CE", "sin_font":     "9C0006",
    "grupo":        "DDEBF7", "grupo_font":   "1F3864",
    "title_bg":     "2E75B6", "title_font":   "FFFFFF",
    "total_bg":     "1F3864",
    "match_exacto": "C6EFCE", "match_exacto_f":"276221",
    "match_difuso": "FFEB9C", "match_difuso_f":"9C5700",
    "huerfano":     "FCE4D6", "huerfano_f":   "843C0C",
}
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_COP  = '#,##0'
FMT_DATE = 'DD/MM/YYYY'
FMT_PCT  = '0.00%'

# ── Lista oficial de sedes (fuente de verdad única) ────────────────────────
SEDES_OFICIALES = [
    "Barranquilla","Buenavista","Centro Mayor","Chipichape",
    "Eden","Envigado","Fabricato","Plaza Imperial",
    "Jardin Plaza","Mercurio","Molinos","Nuestro",
    "Pasto","Puerta del Norte","Santa Marta","Santa Fe",
    "Sincelejo","Parque Alegra","Americas","Cacique",
    "Tesoro","Titan Plaza","Cali","Pereira","Serrezuela",
]

# Prefijos a quitar del campo Tercero (PLAZA/PARQUE/VIVA/LOCAL se conservan)
PREFIJOS_TERCERO = [
    r'^CLIENTES\s+VENTAS\s+',
    r'^CLIENTES\s+VENTA\s+',
    r'^CLIENTES\s+',
]

# Patrones para limpiar sufijos de mes/año en nombres de archivo
_RE_SUFIJO_MES = re.compile(
    r'[_\s]+(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|'
    r'SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE|'
    r'ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)'
    r'([_\s]+\d{2,4})?$',
    re.IGNORECASE
)
_RE_SUFIJO_ANIO = re.compile(
    r'[_\s]+(0[1-9]|1[0-2])[_\s]+\d{4}$|[_\s]+\d{4}$',
    re.IGNORECASE
)

STOPWORDS = {
    'DE','DEL','LA','LAS','LOS','EL','EN','Y','A',
    'MALL','CC','ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
    'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE',
    'THE','OF','VIVA','LOCAL','TIENDA','PARQUE','PLAZA','UNICENTRO',
    'CLIENTES','VENTAS','VENTA',
}
UMBRAL_MATCH = 60


# ════════════════════════════════════════════════════════════════════════════
# 1. COINCIDENCIA DIFUSA
# ════════════════════════════════════════════════════════════════════════════

def _normalizar(s: str) -> str:
    s = s.upper().strip()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def _score_match(nombre_archivo: str, sede_auxiliar: str) -> tuple:
    a = _normalizar(nombre_archivo)
    b = _normalizar(sede_auxiliar)
    if a == b:             return 100.0, 'exacto'
    if a in b:             return 95.0,  'contenido'
    if b in a:             return 90.0,  'contenido-inv'
    wa = set(a.split()) - STOPWORDS
    wb = set(b.split()) - STOPWORDS
    if wa and wb:
        intersect = wa & wb
        score = len(intersect) / max(len(wa), len(wb)) * 85
        if score >= UMBRAL_MATCH:
            return round(score, 1), 'palabras'
    return 0.0, 'sin_match'


def construir_mapa_nombres(nombres_archivo: list, sedes_auxiliar: list) -> dict:
    """
    Asigna cada archivo de datafono a la sede del auxiliar con mayor score.
    Cada sede solo puede usarse una vez (asignación greedy).
    Retorna {nombre_archivo: {sede, score, tipo, confirmado}}
    """
    todos = []
    for nombre in nombres_archivo:
        for sede in sedes_auxiliar:
            score, tipo = _score_match(nombre, sede)
            if score >= UMBRAL_MATCH:
                todos.append((score, nombre, tipo, sede))
    todos.sort(key=lambda x: -x[0])

    mapa = {n: {'sede': None, 'score': 0.0, 'tipo': 'sin_match', 'confirmado': False}
            for n in nombres_archivo}
    sedes_tomadas = set()

    for score, nombre, tipo, sede in todos:
        if mapa[nombre]['sede'] is not None: continue
        if sede in sedes_tomadas:           continue
        mapa[nombre] = {'sede': sede, 'score': score,
                        'tipo': tipo, 'confirmado': score == 100.0}
        sedes_tomadas.add(sede)

    return mapa


# ════════════════════════════════════════════════════════════════════════════
# 2. LECTURA DE ARCHIVOS
# ════════════════════════════════════════════════════════════════════════════

def _limpiar_tercero(tercero: str) -> str:
    """Quita CLIENTES VENTAS / CLIENTES VENTA / CLIENTES del campo Tercero.
    NO quita PLAZA, PARQUE, VIVA, LOCAL — son parte del nombre de la sede."""
    t = str(tercero).strip().upper()
    t = unicodedata.normalize('NFKD', t)
    t = ''.join(c for c in t if not unicodedata.combining(c))
    for patron in PREFIJOS_TERCERO:
        t = re.sub(patron, '', t, flags=re.IGNORECASE)
    return t.strip()


def _extraer_dia_de_nota(nota: str) -> int | None:
    """
    Extrae el día del mes desde la Nota SIN usar la columna Fecha.
    Orden de prioridad:
      1. Fecha completa dd/mm/aaaa o dd-mm-aaaa  → usa el día
      2. Fecha parcial  dd/mm  o  dd-mm           → usa el día
      3. Número al inicio seguido de guion/espacio (formato nuevo)
      4. Cualquier número 1-31 en el texto
    Nunca retorna None si hay un número válido.
    """
    n = str(nota).strip()
    # 1. Fecha completa
    m = re.search(r'\b(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})\b', n)
    if m: return int(m.group(1))
    # 2. Fecha parcial dd/mm
    m = re.search(r'\b(\d{1,2})[/\-](\d{1,2})\b', n)
    if m: return int(m.group(1))
    # 3. Número al inicio (formato nuevo)
    m = re.match(r'^(\d{1,2})[-\s]', n)
    if m: return int(m.group(1))
    # 4. Cualquier número 1-31
    m = re.search(r'\b([1-9]|[12]\d|3[01])\b', n)
    if m: return int(m.group(1))
    return None


def _detectar_formato_fila(nota: str) -> str:
    """
    Detecta si la fila es formato NUEVO o ANTIGUO.
    NUEVO: la Nota empieza con dígito(s) seguido de guion/espacio + DATAFONO
           Y tiene el nombre de la sede después de DATAFONO
    ANTIGUO: la Nota tiene fecha completa/parcial, o solo dice DATAFONO sin nombre de sede
    """
    n = str(nota).strip()
    # Formato nuevo: "2- DATAFONO ENVIGADO" o "10 DATAFONO MOLINOS"
    if re.match(r'^\d{1,2}[-\s]+DATAFONO\s+\S', n, re.IGNORECASE):
        return 'nuevo'
    return 'antiguo'


def leer_auxiliar(path: str) -> pd.DataFrame:
    """
    Lee el Auxiliar de WorldOffice detectando automáticamente el formato por fila:

    FORMATO NUEVO (Enero 2026 en adelante):
      Nota: "2- DATAFONO ENVIGADO"  → día=2, sede='ENVIGADO' (de la Nota)

    FORMATO ANTIGUO (años anteriores):
      Nota: "01/03/2025 DATAFONO"   → día=1 (de la fecha en Nota), sede del campo Tercero
      Nota puede traer fecha completa, parcial, o fecha en medio del texto.
      Tercero: "CLIENTES VENTAS VIVA ENVIGADO" → sede='VIVA ENVIGADO' (prefijo limpiado)

    Nunca usa la columna Fecha del auxiliar para extraer el día.
    """
    raw  = pd.read_excel(path, sheet_name=0, header=None)
    data = raw.iloc[4:].copy()
    data.columns = raw.iloc[3].tolist()
    data = data.reset_index(drop=True)

    # Filtrar solo filas que mencionan DATAFONO
    mask = data['Nota'].astype(str).str.upper().str.contains('DATAFONO', na=False)
    df   = data[mask].copy()

    df['Debitos']  = pd.to_numeric(df['Debitos'],  errors='coerce').fillna(0)
    df['Creditos'] = pd.to_numeric(df['Creditos'], errors='coerce').fillna(0)
    df['Saldo']    = pd.to_numeric(df['Saldo'],    errors='coerce').fillna(0)

    # Detectar formato y extraer Dia_Operacion + Sede fila por fila
    dias, sedes, formatos = [], [], []
    tercero_col = 'Tercero' if 'Tercero' in df.columns else None

    for _, row in df.iterrows():
        nota = str(row['Nota']).strip()
        fmt  = _detectar_formato_fila(nota)
        dia  = _extraer_dia_de_nota(nota)

        if fmt == 'nuevo':
            # Sede viene después de DATAFONO en la misma Nota
            m = re.search(r'DATAFONO\s+(.+?)$', nota, re.IGNORECASE)
            sede = m.group(1).strip().upper() if m else ''
        else:
            # Formato antiguo: sede viene del Tercero
            if tercero_col and pd.notna(row.get(tercero_col)):
                sede = _limpiar_tercero(str(row[tercero_col]))
            else:
                # fallback: intentar extraer de la Nota después de DATAFONO
                m = re.search(r'DATAFONO\s+(.+?)$', nota, re.IGNORECASE)
                sede = m.group(1).strip().upper() if m else ''

        dias.append(dia)
        sedes.append(sede)
        formatos.append(fmt)

    df['Dia_Operacion'] = dias
    df['Sede']          = sedes
    df['_Formato']      = formatos   # 'nuevo' | 'antiguo' — útil para auditoría
    return df


def sedes_disponibles(path: str) -> list:
    """
    Retorna las sedes únicas tal como existen en la columna Sede del auxiliar.
    Los valores ya pasaron por _limpiar_tercero en leer_auxiliar, así que
    están limpios de prefijos (CLIENTES VENTAS, etc.) pero mantienen su forma
    real (ej: 'VIVA ENVIGADO', 'EL EDEN', 'LOS MOLINOS').
    Estos valores son los que deben usarse para filtrar df['Sede'] en el cruce.
    """
    df  = leer_auxiliar(path)
    raw = df['Sede'].dropna().unique().tolist()
    return sorted(s for s in raw if s and str(s).strip())


def _detectar_encabezado(path, sheet_name):
    probe = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=10)
    for i in range(8):
        vals = probe.iloc[i].astype(str).str.strip().tolist()
        if 'Fecha Vale' in vals and 'Valor Neto' in vals:
            return i
    return None


def _leer_hoja(path, sheet_name):
    hr = _detectar_encabezado(path, sheet_name)
    if hr is None: return None
    df = pd.read_excel(path, sheet_name=sheet_name, header=hr)
    df.columns = [str(c).strip() for c in df.columns]
    if not all(c in df.columns for c in ['Fecha Vale','Fecha de Abono','Bol. Ruta','Valor Neto']):
        return None
    df['Fecha Vale']     = pd.to_datetime(df['Fecha Vale'],     errors='coerce')
    df['Fecha de Abono'] = pd.to_datetime(df['Fecha de Abono'], errors='coerce')
    df['Bol. Ruta']      = df['Bol. Ruta'].astype(str).str.strip()
    df['Valor Neto']     = pd.to_numeric(df['Valor Neto'], errors='coerce').fillna(0)
    for col in ['Valor Comisión','Ret. Fuente','Ret. IVA','Ret. ICA','Valor Consumo']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['_Hoja'] = sheet_name
    return df.dropna(subset=['Fecha Vale','Valor Neto'])


def leer_datafono(path: str) -> pd.DataFrame:
    xl = pd.ExcelFile(path)
    hojas_ok = []
    for sheet in xl.sheet_names:
        df = _leer_hoja(path, sheet)
        if df is not None and len(df) > 0:
            hojas_ok.append(df)
    if not hojas_ok:
        raise ValueError(f"Sin datos válidos en {path}")
    # Anti-duplicado por totales diarios
    unicas = [hojas_ok[0]]
    ref = hojas_ok[0].groupby(hojas_ok[0]['Fecha Vale'].dt.day)['Valor Neto'].sum().to_dict()
    for candidata in hojas_ok[1:]:
        tots = candidata.groupby(candidata['Fecha Vale'].dt.day)['Valor Neto'].sum().to_dict()
        dias_comunes = set(ref.keys()) & set(tots.keys())
        if dias_comunes:
            iguales = sum(1 for d in dias_comunes if abs(ref[d]-tots[d]) < 1)
            if iguales / len(dias_comunes) > 0.8:
                continue
        unicas.append(candidata)
    return pd.concat(unicas, ignore_index=True)


_RE_SUFIJO_MES = re.compile(
    r'[_\s]+(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|'
    r'SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE|'
    r'ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)'
    r'[_\s]*\d{0,4}$',
    re.IGNORECASE,
)
_RE_SUFIJO_ANIO = re.compile(r'[_\s]+\d{4}$')


def _normalizar_nombre_archivo(path: str) -> str:
    """
    Extrae el nombre de sede desde el nombre del archivo de datafono.
    Quita:  prefijo 'datafono_'  +  sufijos de mes/año
    Ejemplos:
      datafono_ENVIGADO_ENERO_2025.xlsx  → ENVIGADO
      datafono_AMERICAS_ENE_2025.xlsx    → AMERICAS
      datafono_MOLINOS_01_2025.xlsx      → MOLINOS
      datafono_ENVIGADO.xlsx             → ENVIGADO
    """
    nombre = _extraer_nombre_sede_de_archivo(
        os.path.splitext(os.path.basename(path))[0]
    )
    nombre = _RE_SUFIJO_MES.sub('', nombre).strip('_').strip()
    nombre = _RE_SUFIJO_ANIO.sub('', nombre).strip('_').strip()
    return nombre

def _extraer_nombre_sede_de_archivo(nombre_raw: str) -> str:
    """
    Extrae el nombre útil de sede desde el nombre del archivo de datafono.
    Maneja múltiples convenciones de nomenclatura:
      - 'datafono_ENVIGADO'              → 'ENVIGADO'
      - 'MOVIMIENTOS DATAFONO ENVIGADO'  → 'ENVIGADO'
      - 'REDEBAN DATAFONO VIVA ENVIGADO' → 'VIVA ENVIGADO'
      - 'ENVIGADO'                       → 'ENVIGADO'
      - 'datafono_VIVA ENVIGADO'         → 'VIVA ENVIGADO'
    Regla: si contiene la palabra DATAFONO, usar lo que viene DESPUÉS.
    Si no, quitar prefijo 'datafono_' y usar el resto.
    """
    n = nombre_raw.strip().upper()
    # Caso 1: contiene la palabra DATAFONO → extraer lo que viene después
    m = re.search(r'\bDATAFONO\b\s+(.+)', n, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    # Caso 2: prefijo datafono_ (con guion bajo)
    n = re.sub(r'^DATAFONO[_\s]+', '', n, flags=re.IGNORECASE)
    # Caso 3: prefijos operativos conocidos que no aportan al nombre de sede
    for prefijo in [r'^MOVIMIENTOS\s+', r'^REPORTE\s+', r'^REDEBAN\s+',
                    r'^CREDIBANCO\s+', r'^INFORME\s+', r'^DETALLE\s+']:
        n = re.sub(prefijo, '', n, flags=re.IGNORECASE)
    return n.strip()

def cargar_multiples_datafonos(archivos: list) -> tuple:
    """Retorna (df_unificado, info_archivos{nombre: {path,filas}})"""
    frames, info = [], {}
    for arch in archivos:
        nombre = _normalizar_nombre_archivo(arch)
        try:
            df = leer_datafono(arch)
            df['_Nombre_Archivo'] = nombre
            frames.append(df)
            info[nombre] = {'path': arch, 'filas': len(df)}
        except Exception as e:
            info[nombre] = {'path': arch, 'filas': 0, 'error': str(e)}
    df_total = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return df_total, info


# ════════════════════════════════════════════════════════════════════════════
# 3. AGRUPACIÓN Y CRUCE
# ════════════════════════════════════════════════════════════════════════════

def agrupar_datafono_por_dia_vale(df_dat: pd.DataFrame) -> tuple:
    df = df_dat.copy()
    df['_Dia_Vale']  = df['Fecha Vale'].dt.day
    df['_Mes_Vale']  = df['Fecha Vale'].dt.month
    df['_Year_Vale'] = df['Fecha Vale'].dt.year

    agg = {'Valor Neto': 'sum'}
    for col in ['Valor Comisión','Ret. Fuente','Ret. IVA','Ret. ICA']:
        if col in df.columns: agg[col] = 'sum'

    por_dia = df.groupby(
        ['_Nombre_Archivo','_Year_Vale','_Mes_Vale','_Dia_Vale'], dropna=False
    ).agg(agg).reset_index()

    por_bolruta = df.groupby(
        ['_Nombre_Archivo','_Year_Vale','_Mes_Vale','_Dia_Vale','Fecha de Abono','Bol. Ruta'],
        dropna=False
    ).agg({'Valor Neto': 'sum'}).reset_index()

    return por_dia, por_bolruta


def cruzar_auxiliar_datafono(df_aux, por_dia, por_bolruta,
                              nombre_archivo: str,
                              year: int = 2026, mes: int = 1) -> list:
    resultados = []
    for _, ra in df_aux.iterrows():
        dia   = ra['Dia_Operacion']
        v_aux = float(ra['Debitos'])
        if pd.isna(dia):
            resultados.append({'row_aux': ra, 'candidatos_bolruta': pd.DataFrame(),
                                'suma_grupos': 0, 'diferencia': v_aux,
                                'estado': 'SIN_DIA', 'valor_aux': v_aux})
            continue
        dia = int(dia)
        match_dia = por_dia[
            (por_dia['_Nombre_Archivo'].str.upper() == nombre_archivo.upper()) &
            (por_dia['_Dia_Vale']  == dia)  &
            (por_dia['_Year_Vale'] == year) &
            (por_dia['_Mes_Vale']  == mes)
        ]
        s_gr  = float(match_dia['Valor Neto'].sum()) if len(match_dia) > 0 else 0.0
        dif   = v_aux - s_gr
        cands = por_bolruta[
            (por_bolruta['_Nombre_Archivo'].str.upper() == nombre_archivo.upper()) &
            (por_bolruta['_Dia_Vale']  == dia)  &
            (por_bolruta['_Year_Vale'] == year) &
            (por_bolruta['_Mes_Vale']  == mes)
        ].copy()

        if len(match_dia) == 0:      estado = 'SIN_MATCH'
        elif abs(dif) < 1:           estado = 'CUADRA'
        elif abs(dif) <= v_aux*0.05: estado = 'DIF_MENOR'
        else:                        estado = 'DIFERENCIA'

        resultados.append({'row_aux': ra, 'candidatos_bolruta': cands,
                            'suma_grupos': s_gr, 'diferencia': dif,
                            'estado': estado, 'valor_aux': v_aux})
    return resultados


def agrupar_datafono_por_abono(df_dat: pd.DataFrame) -> tuple:
    """
    Agrupa el datafono por Fecha de ABONO (en lugar de Fecha Vale).
    Usado exclusivamente en el modo de conciliación por Fecha de Abono
    para períodos históricos donde WorldOffice registró por fecha de abono.

    Retorna:
      por_abono   — agrupado por (archivo, año_abono, mes_abono, dia_abono)
      por_bolruta — detalle por bol_ruta con Fecha Vale como referencia informativa
    """
    df = df_dat.copy()
    df['_Dia_Abono']  = df['Fecha de Abono'].dt.day
    df['_Mes_Abono']  = df['Fecha de Abono'].dt.month
    df['_Year_Abono'] = df['Fecha de Abono'].dt.year

    agg = {'Valor Neto': 'sum'}
    for col in ['Valor Comisión', 'Ret. Fuente', 'Ret. IVA', 'Ret. ICA']:
        if col in df.columns:
            agg[col] = 'sum'

    por_abono = df.groupby(
        ['_Nombre_Archivo', '_Year_Abono', '_Mes_Abono', '_Dia_Abono'], dropna=False
    ).agg(agg).reset_index()

    # Detalle por bol_ruta — Fecha Vale es solo informativa en este modo
    por_bolruta = df.groupby(
        ['_Nombre_Archivo', '_Year_Abono', '_Mes_Abono', '_Dia_Abono', 'Fecha Vale', 'Bol. Ruta'],
        dropna=False
    ).agg({'Valor Neto': 'sum'}).reset_index()

    return por_abono, por_bolruta


def cruzar_auxiliar_datafono_por_abono(df_aux, por_abono, por_bolruta,
                                        nombre_archivo: str,
                                        year: int = 2025, mes: int = 1) -> list:
    """
    Cruza el auxiliar contra el datafono usando Fecha de Abono como clave.

    Regla: auxiliar.Dia_Operacion == datafono.Fecha_Abono.day
           para el mismo mes y año del período conciliado.

    Registros cuya nota indica un mes distinto al período (ej: nota '01/02/2025'
    en el auxiliar de enero) quedan como SIN_MATCH — son registros de cierre de
    mes que contabilidad incluyó anticipadamente y no pertenecen al período.

    ADVERTENCIA NORMATIVA: Este modo es contablemente incorrecto según NIIF PYMES
    (principio de devengo). Usar únicamente para conciliación de períodos históricos
    donde el auxiliar fue registrado por Fecha de Abono. Requiere nota en el informe.
    """
    resultados = []
    for _, ra in df_aux.iterrows():
        dia   = ra['Dia_Operacion']
        v_aux = float(ra['Debitos'])

        if pd.isna(dia):
            resultados.append({'row_aux': ra, 'candidatos_bolruta': pd.DataFrame(),
                                'suma_grupos': 0, 'diferencia': v_aux,
                                'estado': 'SIN_DIA', 'valor_aux': v_aux,
                                'modo_cruce': 'FECHA_ABONO'})
            continue

        dia = int(dia)

        match_abono = por_abono[
            (por_abono['_Nombre_Archivo'].str.upper() == nombre_archivo.upper()) &
            (por_abono['_Dia_Abono']  == dia)  &
            (por_abono['_Year_Abono'] == year) &
            (por_abono['_Mes_Abono']  == mes)
        ]

        s_gr = float(match_abono['Valor Neto'].sum()) if len(match_abono) > 0 else 0.0
        dif  = v_aux - s_gr

        cands = por_bolruta[
            (por_bolruta['_Nombre_Archivo'].str.upper() == nombre_archivo.upper()) &
            (por_bolruta['_Dia_Abono']  == dia)  &
            (por_bolruta['_Year_Abono'] == year) &
            (por_bolruta['_Mes_Abono']  == mes)
        ].copy()

        if len(match_abono) == 0:    estado = 'SIN_MATCH'
        elif abs(dif) < 1:           estado = 'CUADRA'
        elif abs(dif) <= v_aux*0.05: estado = 'DIF_MENOR'
        else:                        estado = 'DIFERENCIA'

        resultados.append({'row_aux': ra, 'candidatos_bolruta': cands,
                            'suma_grupos': s_gr, 'diferencia': dif,
                            'estado': estado, 'valor_aux': v_aux,
                            'modo_cruce': 'FECHA_ABONO'})
    return resultados


# ════════════════════════════════════════════════════════════════════════════
# 4. UTILIDADES EXCEL
# ════════════════════════════════════════════════════════════════════════════

def _sc(cell, bg=None, fc="000000", bold=False, fmt=None, ha="left", sz=10):
    if bg: cell.fill = PatternFill("solid", start_color=bg, fgColor=bg)
    cell.font      = Font(color=fc, bold=bold, name="Arial", size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=False)
    cell.border    = BORDER
    if fmt: cell.number_format = fmt


def _titulo(ws, texto, subtexto="", ncols=15):
    span = f"A1:{get_column_letter(ncols)}1"
    ws.merge_cells(span)
    t = ws["A1"]
    t.value = texto
    t.fill  = PatternFill("solid", start_color=COLOR["title_bg"])
    t.font  = Font(color="FFFFFF", bold=True, name="Arial", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    if subtexto:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        g = ws["A2"]
        g.value = subtexto
        g.font  = Font(color="595959", italic=True, name="Arial", size=8)
        g.alignment = Alignment(horizontal="center")


# ════════════════════════════════════════════════════════════════════════════
# 5. HOJA DE DETALLE (reutilizable)
# ════════════════════════════════════════════════════════════════════════════

def _hoja_detalle(wb, resultados, sede, periodo, ws_name="DETALLE_CONCILIACION"):
    ws = wb.create_sheet(ws_name)
    _titulo(ws, f"DETALLE — {sede.upper()} — {periodo.upper()}",
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
            "Cruce: Día(Nota) = Día(Fecha Vale)  |  Normativa NIIF PYMES / DIAN")

    HEADERS = ["Tipo Fila","Día","Nota Auxiliar","Doc Num",
               "Valor Auxiliar ($)","Sum Valor Neto\nDatafono ($)",
               "Diferencia ($)","% Dif","Estado",
               "Fecha\nAbono","Bol_Ruta",
               "Valor Neto\nBol_Ruta ($)","Comisión +\nRet ($)",
               "# Bol_\nRutas","Observación"]
    widths = [15,6,42,18,18,19,14,8,14,12,14,18,14,9,36]

    for ci,(h,w) in enumerate(zip(HEADERS,widths),1):
        cell = ws.cell(row=3, column=ci, value=h)
        _sc(cell, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 36

    row_n = 4
    total_aux = total_df = 0

    for res in resultados:
        ra,estado = res['row_aux'], res['estado']
        v_aux,s_gr,dif = res['valor_aux'], res['suma_grupos'], res['diferencia']
        cands = res['candidatos_bolruta']
        dia   = int(ra['Dia_Operacion']) if pd.notna(ra['Dia_Operacion']) else None
        total_aux += v_aux; total_df += s_gr

        if estado == 'CUADRA':
            bg,fc,obs = COLOR["cuadra"],    COLOR["cuadra_font"],  "✔ Cuadra exacto"
        elif estado == 'DIF_MENOR':
            bg,fc,obs = COLOR["dif_menor"], COLOR["dif_font"],     "⚠ Dif ≤5% — comisión/retención/redondeo"
        else:
            bg,fc,obs = COLOR["diferencia"],COLOR["sin_font"],     (
                "❌ Sin coincidencia en datafono" if estado in ('SIN_MATCH','SIN_DIA')
                else "❌ Diferencia significativa — revisar")

        pct  = (dif / v_aux) if v_aux != 0 else 0
        vals = ["AUXILIAR",dia,str(ra['Nota']).strip(),str(ra['Doc Num']).strip(),
                v_aux, s_gr if s_gr>0 else None, dif if s_gr>0 else None,
                pct if s_gr>0 else None, estado,
                None,None,None,None, len(cands) if cands is not None and len(cands)>0 else None, obs]
        fmts   = [None,None,None,None,FMT_COP,FMT_COP,FMT_COP,FMT_PCT,None,None,None,None,None,None,None]
        aligns = ["center","center","left","left","right","right","right","right",
                  "center","left","left","right","right","center","left"]
        for ci,(v,fmt,ha) in enumerate(zip(vals,fmts,aligns),1):
            cell = ws.cell(row=row_n, column=ci, value=v)
            _sc(cell, bg=bg, fc=fc, bold=(ci in [1,5,6,7]), fmt=fmt, ha=ha)
        ws.row_dimensions[row_n].height = 18
        row_n += 1

        if cands is not None:
            for _,gr in cands.iterrows():
                com_ret = sum(float(gr[c]) for c in
                              ['Valor Comisión','Ret. Fuente','Ret. IVA','Ret. ICA']
                              if c in gr.index and pd.notna(gr[c]))
                # Fecha de Abono puede no existir en modo FECHA_ABONO (el bolruta
                # se agrupa por Fecha Vale como referencia informativa en ese modo)
                fa = gr['Fecha de Abono'] if 'Fecha de Abono' in gr.index else None
                fv = gr['Fecha Vale']     if 'Fecha Vale'     in gr.index else None
                try:
                    dias_ab = (fa - fv).days if (fa is not None and fv is not None
                                                 and pd.notna(fa) and pd.notna(fv)) else '?'
                except Exception:
                    dias_ab = '?'
                fa_display = fa.date() if (fa is not None and pd.notna(fa)) else (
                             fv.date() if (fv is not None and pd.notna(fv)) else None)
                vals_gr = ["GRUPO_BOL_RUTA",None,None,None,None,None,None,None,None,
                            fa_display,
                            str(gr['Bol. Ruta']), gr['Valor Neto'],
                            com_ret if com_ret>0 else None, None, f"Abono D+{dias_ab}"]
                fmts_gr   = [None]*9+[FMT_DATE,None,FMT_COP,FMT_COP,None,None]
                aligns_gr = ["center"]+["left"]*9+["left","right","right","center","left"]
                for ci,(v,fmt,ha) in enumerate(zip(vals_gr,fmts_gr,aligns_gr),1):
                    cell = ws.cell(row=row_n, column=ci, value=v)
                    _sc(cell, bg=COLOR["grupo"], fc=COLOR["grupo_font"], fmt=fmt, ha=ha)
                ws.row_dimensions[row_n].height = 15
                row_n += 1

    total_dif = total_aux - total_df
    tots = ["","","TOTAL CONCILIADO","",total_aux,total_df,total_dif,"","","","","","","",""]
    for ci,v in enumerate(tots,1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        _sc(cell, bg=COLOR["total_bg"], fc="FFFFFF", bold=True,
            fmt=FMT_COP if ci in [5,6,7] else None,
            ha="right" if ci in [5,6,7] else "center")
    ws.row_dimensions[row_n].height = 20
    ws.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════════════════
# 6. HOJA RESUMEN (individual)
# ════════════════════════════════════════════════════════════════════════════

def _hoja_resumen_individual(wb, resultados, sede, periodo, info_match=None):
    ws = wb.create_sheet("RESUMEN", 0)
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22

    cuadra    = sum(1 for r in resultados if r['estado'] == 'CUADRA')
    dif_menor = sum(1 for r in resultados if r['estado'] == 'DIF_MENOR')
    diferencia= sum(1 for r in resultados if r['estado'] == 'DIFERENCIA')
    sin_match = sum(1 for r in resultados if r['estado'] in ('SIN_MATCH','SIN_DIA'))
    total_aux = sum(r['valor_aux']   for r in resultados)
    total_df  = sum(r['suma_grupos'] for r in resultados)
    total_dif = total_aux - total_df
    total_reg = len(resultados)
    pct_ok    = (cuadra / total_reg * 100) if total_reg > 0 else 0

    _titulo(ws, f"CONCILIACIÓN DATAFONO — {sede.upper()} — {periodo.upper()}",
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Normativa NIIF PYMES / DIAN Colombia",
            ncols=3)

    def put(r,c,v,bg=None,bold=False,fmt=None,ha="left"):
        cell = ws.cell(row=r, column=c, value=v)
        _sc(cell, bg=bg, bold=bold, fmt=fmt, ha=ha)

    start = 4
    if info_match and info_match.get('tipo') != 'exacto':
        ws.merge_cells("A3:C3")
        m = ws["A3"]
        m.value = (f"⚠ Match difuso: archivo '{info_match['nombre_archivo']}' → sede '{sede}' "
                   f"(score {info_match['score']:.0f}/100 — {info_match['tipo']}) — confirmar con auxiliar contable")
        m.fill  = PatternFill("solid", start_color=COLOR["match_difuso"])
        m.font  = Font(color=COLOR["match_difuso_f"], bold=True, name="Arial", size=9)
        m.alignment = Alignment(horizontal="left", wrap_text=True)
        ws.row_dimensions[3].height = 22

    r = start
    put(r,1,"INDICADOR", bg=COLOR["header_bg"],bold=True,ha="center")
    put(r,2,"CANTIDAD",  bg=COLOR["header_bg"],bold=True,ha="center")
    put(r,3,"VALOR ($)", bg=COLOR["header_bg"],bold=True,ha="center")
    ws.row_dimensions[r].height = 20

    stats = [
        ("Total registros auxiliar (DATAFONO)",            total_reg,   total_aux, None),
        ("✔  Cuadran exacto (diferencia < $1)",            cuadra,      None,      COLOR["cuadra"]),
        ("⚠  Diferencia menor ≤5% (comisión / retención)", dif_menor,   None,      COLOR["dif_menor"]),
        ("❌  Diferencia significativa — revisar",          diferencia,  None,      COLOR["diferencia"]),
        ("❌  Sin coincidencia en datafono",                sin_match,   None,      COLOR["diferencia"]),
        (f"📊  % registros cuadrados",                     f"{pct_ok:.1f}%", None, None),
    ]
    for i,(lbl,cant,val,bg) in enumerate(stats,1):
        put(r+i,1,lbl, bg=bg)
        put(r+i,2,cant,bg=bg,ha="center")
        put(r+i,3,val if val else "",bg=bg,fmt=FMT_COP if val else None,ha="right")
        ws.row_dimensions[r+i].height = 18

    r2 = r + len(stats) + 2
    ws.merge_cells(f"A{r2}:C{r2}")
    cf = ws[f"A{r2}"]
    cf.value="RESUMEN FINANCIERO"
    cf.fill = PatternFill("solid",start_color=COLOR["header_bg"])
    cf.font = Font(color="FFFFFF",bold=True,name="Arial",size=11)
    cf.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[r2].height = 22

    bg_dif = (COLOR["cuadra"] if abs(total_dif)<1
              else COLOR["dif_menor"] if abs(total_dif)<total_aux*0.05
              else COLOR["diferencia"])
    for i,(lbl,val,bg) in enumerate([
        ("Total registrado en auxiliar WorldOffice ($)",          total_aux, None),
        ("Total valor neto datafono — Redeban / Credibanco ($)",  total_df,  None),
        ("Diferencia neta ($)",                                   total_dif, bg_dif),
    ],1):
        put(r2+i,1,lbl,bg=bg)
        put(r2+i,2,"",bg=bg)
        put(r2+i,3,val,bg=bg,fmt=FMT_COP,ha="right",bold=(i==3))
        ws.row_dimensions[r2+i].height = 18

    r3 = r2 + 5
    ws.merge_cells(f"A{r3}:C{r3}")
    nota = ws.cell(row=r3, column=1)
    nota.value = ("CRITERIO CONTABLE:  Reconocimiento → Fecha Vale  |  "
                  "Conciliación bancaria → Fecha de Abono  |  Valor conciliable → Valor Neto")
    nota.font      = Font(color="595959",italic=True,name="Arial",size=8)
    nota.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r3].height = 28


# ════════════════════════════════════════════════════════════════════════════
# 7. HOJA RESUMEN EJECUTIVO (todos)
# ════════════════════════════════════════════════════════════════════════════

def _hoja_resumen_ejecutivo(wb, resumen_sedes, huerfanos, periodo):
    ws = wb.create_sheet("RESUMEN_EJECUTIVO", 0)
    _titulo(ws, f"RESUMEN EJECUTIVO — TODAS LAS SEDES — {periodo.upper()}",
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
            "Normativa NIIF PYMES / DIAN Colombia", ncols=12)

    HEADERS = ["Sede Auxiliar","Archivo Datafono","Tipo Match","Score",
               "Registros","✔ Cuadran","⚠ Dif Menor","❌ Revisar",
               "Total Auxiliar ($)","Total Datafono ($)","Diferencia ($)","% Cuadrado"]
    widths  = [26,22,14,8,11,11,12,11,20,20,18,13]

    for ci,(h,w) in enumerate(zip(HEADERS,widths),1):
        cell = ws.cell(row=3, column=ci, value=h)
        _sc(cell,bg=COLOR["header_bg"],fc=COLOR["header_font"],bold=True,ha="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    row_n = 4
    for res in resumen_sedes:
        pct   = (res['cuadra']/res['total_reg']*100) if res['total_reg']>0 else 0
        tipo  = res.get('tipo_match','—')
        score = res.get('score',0)
        if res['revisar']==0:
            bg,fc = (COLOR["cuadra"],COLOR["cuadra_font"]) if tipo=='exacto' else (COLOR["dif_menor"],COLOR["dif_font"])
        else:
            bg,fc = COLOR["diferencia"],COLOR["sin_font"]

        vals = [res['sede'], res.get('nombre_df','—'), tipo, score,
                res['total_reg'],res['cuadra'],res['dif_menor'],res['revisar'],
                res['total_aux'],res['total_df'],res['diferencia'], pct/100]
        fmts = [None,None,None,None,None,None,None,None,FMT_COP,FMT_COP,FMT_COP,FMT_PCT]
        aligns = ["left","left","center","center","center","center","center","center",
                  "right","right","right","right"]
        for ci,(v,fmt,ha) in enumerate(zip(vals,fmts,aligns),1):
            cell = ws.cell(row=row_n, column=ci, value=v)
            _sc(cell,bg=bg,fc=fc,fmt=fmt,ha=ha,bold=(ci==1))
        ws.row_dimensions[row_n].height = 18
        row_n += 1

    if huerfanos:
        row_n += 1
        ws.merge_cells(f"A{row_n}:L{row_n}")
        h = ws.cell(row=row_n, column=1,
                    value=f"⚠  {len(huerfanos)} ARCHIVO(S) DATAFONO SIN SEDE CORRESPONDIENTE EN EL AUXILIAR")
        h.fill = PatternFill("solid", start_color=COLOR["huerfano"])
        h.font = Font(color=COLOR["huerfano_f"], bold=True, name="Arial", size=10)
        h.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row_n].height = 20
        row_n += 1
        for nombre in huerfanos:
            ws.merge_cells(f"A{row_n}:L{row_n}")
            cell = ws.cell(row=row_n, column=1,
                           value=f"     ❌  {nombre}  — Renombrar el archivo o verificar que la sede existe en el auxiliar")
            _sc(cell, bg=COLOR["huerfano"], fc=COLOR["huerfano_f"])
            ws.row_dimensions[row_n].height = 16
            row_n += 1

    row_n += 1
    tots = ["TOTAL GENERAL","","","",
            sum(r['total_reg'] for r in resumen_sedes),
            sum(r['cuadra']    for r in resumen_sedes),
            sum(r['dif_menor'] for r in resumen_sedes),
            sum(r['revisar']   for r in resumen_sedes),
            sum(r['total_aux'] for r in resumen_sedes),
            sum(r['total_df']  for r in resumen_sedes),
            sum(r['diferencia']for r in resumen_sedes), ""]
    fmts_t = [None,None,None,None,None,None,None,None,FMT_COP,FMT_COP,FMT_COP,None]
    for ci,(v,fmt) in enumerate(zip(tots,fmts_t),1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        _sc(cell,bg=COLOR["total_bg"],fc="FFFFFF",bold=True,fmt=fmt,
            ha="right" if ci>8 else "center" if ci>1 else "left")
    ws.row_dimensions[row_n].height = 22
    ws.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════════════════
# 8. HOJA MAPA DE NOMBRES
# ════════════════════════════════════════════════════════════════════════════

def _hoja_mapa_nombres(wb, mapa_nombres, huerfanos):
    ws = wb.create_sheet("MAPA_NOMBRES")
    _titulo(ws, "MAPA DE NOMBRES — COINCIDENCIA DIFUSA",
            "⚠  Revisar y confirmar manualmente los matches no exactos antes de aprobar el reporte",
            ncols=6)

    HEADERS = ["Archivo Datafono","Sede Auxiliar Asignada","Tipo de Match",
               "Score (0-100)","¿Confirmar?","Observación"]
    widths  = [28,32,18,14,22,42]
    for ci,(h,w) in enumerate(zip(HEADERS,widths),1):
        cell = ws.cell(row=3, column=ci, value=h)
        _sc(cell,bg=COLOR["header_bg"],fc=COLOR["header_font"],bold=True,ha="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    row_n = 4
    for nombre, info in sorted(mapa_nombres.items()):
        sede  = info['sede'] or "⚠ SIN ASIGNAR"
        score = info['score']
        tipo  = info['tipo']
        conf  = info['confirmado']

        if sede == "⚠ SIN ASIGNAR":
            bg,fc = COLOR["huerfano"],     COLOR["huerfano_f"]
            req   = "— Huérfano"
            obs   = "Ninguna sede del auxiliar coincide con este nombre de archivo"
        elif conf:
            bg,fc = COLOR["match_exacto"], COLOR["match_exacto_f"]
            req   = "No — match exacto"
            obs   = "Nombres idénticos. No requiere revisión."
        else:
            bg,fc = COLOR["match_difuso"], COLOR["match_difuso_f"]
            req   = "⚠ SÍ — confirmar"
            obs   = {
                'contenido':     f"'{nombre}' está contenido dentro de '{sede}'",
                'contenido-inv': f"'{sede}' está contenido dentro de '{nombre}'",
                'palabras':      f"Palabras significativas en común (score {score:.0f}/100)",
            }.get(tipo, "Coincidencia parcial — revisar manualmente")

        vals = [nombre, sede, tipo, score if score>0 else "—", req, obs]
        for ci,v in enumerate(vals,1):
            cell = ws.cell(row=row_n, column=ci, value=v)
            _sc(cell,bg=bg,fc=fc,bold=(ci==1),ha="center" if ci in [3,4] else "left")
        ws.row_dimensions[row_n].height = 18
        row_n += 1

    ws.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════════════════
# 9. HOJA LOG DE AUDITORÍA
# ════════════════════════════════════════════════════════════════════════════

def _hoja_auditoria(wb, sede, periodo, info_match=None, es_consolidado=False,
                    mapa_nombres=None, huerfanos=None, modo_fecha='FECHA_VALE'):
    ws = wb.create_sheet("LOG_AUDITORIA")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 66
    _titulo(ws, "LOG DE AUDITORÍA — INMUTABLE", ncols=2)

    sedes_exactas = sum(1 for v in (mapa_nombres or {}).values() if v['confirmado'])
    sedes_difusas = sum(1 for v in (mapa_nombres or {}).values()
                        if not v['confirmado'] and v['sede'])

    match_txt = "Exacto (100/100)" if not info_match else (
        f"{info_match.get('nombre_archivo','')} → {sede} | "
        f"score={info_match['score']:.0f} | tipo={info_match['tipo']}")

    if modo_fecha == 'FECHA_ABONO':
        logica_cruce  = "Nota.Dia_Operacion == Fecha_Abono.day (mismo mes y año) — MODO HISTÓRICO"
        agrupacion_df = "Por día Fecha de Abono → SUM(Valor Neto)"
        advertencia   = ("⚠ MODO FECHA ABONO — Criterio no estándar NIIF. Usado para períodos "
                         "históricos donde el auxiliar fue registrado por fecha de abono. "
                         "No usar para períodos corrientes.")
    else:
        logica_cruce  = "Nota.Dia_Operacion == Fecha_Vale.day (mismo mes y año)"
        agrupacion_df = "Por día Fecha Vale → SUM(Valor Neto)"
        advertencia   = None

    registros = [
        ("Fecha / Hora proceso",          datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ("Sede / Modo",                   sede if not es_consolidado else "Consolidado — todas las sedes"),
        ("Período",                       periodo),
        ("Versión motor",                 "1.5.1"),
        ("Modo de cruce de fechas",       modo_fecha),
        ("Normativa",                     "NIIF para PYMES (Decreto 2420 de 2015) / DIAN Colombia"),
        ("Lógica de cruce",               logica_cruce),
        ("Matching de nombres",           match_txt if not es_consolidado else
                                          f"Exactos: {sedes_exactas} | Difusos: {sedes_difusas} | Huérfanos: {len(huerfanos or [])}"),
        ("Umbral match difuso",           f"{UMBRAL_MATCH}/100"),
        ("Agrupación datafono",           agrupacion_df),
        ("Detalle bancario",              "Por (Fecha Abono + Bol_Ruta) → SUM(Valor Neto)"),
        ("Tolerancia diferencia menor",   "≤ 5% del valor auxiliar"),
        ("Anti-duplicado hojas xlsx",     "Si >80% días con totales idénticos → usa hoja principal"),
        ("Archivos originales",           "NO modificados — motor opera sobre copias en memoria"),
        ("Auditoría externa",             "KPMG Ltda. — Davivienda Cuenta Corriente 2346"),
    ]
    if advertencia:
        registros.append(("⚠ ADVERTENCIA NORMATIVA", advertencia))
    if es_consolidado and huerfanos:
        registros.append(("Huérfanos — lista completa", " | ".join(huerfanos)))

    for i,(k,v) in enumerate(registros,2):
        ck = ws.cell(row=i, column=1, value=k)
        cv = ws.cell(row=i, column=2, value=v)
        _sc(ck, bg="F2F2F2", bold=True)
        _sc(cv)
        ws.row_dimensions[i].height = 16


# ════════════════════════════════════════════════════════════════════════════
# 10. FUNCIONES PÚBLICAS PRINCIPALES
# ════════════════════════════════════════════════════════════════════════════



# ════════════════════════════════════════════════════════════════════════════
# HOJA PENDIENTES
# ════════════════════════════════════════════════════════════════════════════

def _hoja_pendientes(wb, resultados, por_dia, por_bolruta,
                     nombre_archivo, year, mes, sede):
    """
    Hoja PENDIENTES con dos secciones:
    A) Registros del auxiliar que NO cuadraron (SIN_MATCH, DIFERENCIA, SIN_DIA)
    B) Días del datafono que NO tienen contrapartida en el auxiliar
    """
    ws = wb.create_sheet("PENDIENTES")
    _titulo(ws, f"PENDIENTES — {sede.upper()} — Registros sin conciliar",
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
            "Revisar cada ítem antes de cerrar el período", ncols=10)

    # ── A. Auxiliar sin cuadrar ──────────────────────────────────────────────
    pendientes_aux = [r for r in resultados
                      if r['estado'] in ('SIN_MATCH','DIFERENCIA','SIN_DIA')]

    ws.merge_cells("A3:J3")
    h = ws["A3"]
    h.value = f"A)  REGISTRADOS EN AUXILIAR SIN CONCILIAR ({len(pendientes_aux)} registros)"
    h.fill  = PatternFill("solid", start_color=COLOR["diferencia"])
    h.font  = Font(color=COLOR["sin_font"], bold=True, name="Arial", size=10)
    h.alignment = Alignment(horizontal="left")
    ws.row_dimensions[3].height = 20

    HEADERS_A = ["Día","Nota Auxiliar","Doc Num","Valor Auxiliar ($)",
                 "Sum Datafono ($)","Diferencia ($)","Estado","Motivo"]
    widths_a   = [6, 44, 18, 20, 20, 18, 14, 36]
    for ci,(h,w) in enumerate(zip(HEADERS_A,widths_a),1):
        cell = ws.cell(row=4, column=ci, value=h)
        _sc(cell, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 18

    row_n = 5
    total_aux_pdte = 0
    for res in pendientes_aux:
        ra    = res['row_aux']
        dia   = int(ra['Dia_Operacion']) if pd.notna(ra['Dia_Operacion']) else '?'
        v_aux = res['valor_aux']
        s_gr  = res['suma_grupos']
        dif   = res['diferencia']
        total_aux_pdte += v_aux

        motivo = {
            'SIN_MATCH':  'No se encontró ningún grupo en el datafono para este día',
            'DIFERENCIA': 'Diferencia > 5% — valor en datafono no corresponde al auxiliar',
            'SIN_DIA':    'No se pudo extraer el día de la Nota — revisar formato',
        }.get(res['estado'], res['estado'])

        vals = [dia, str(ra['Nota']).strip(), str(ra['Doc Num']).strip(),
                v_aux, s_gr if s_gr > 0 else None,
                dif   if s_gr > 0 else None,
                res['estado'], motivo]
        fmts = [None,None,None,FMT_COP,FMT_COP,FMT_COP,None,None]
        for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
            cell = ws.cell(row=row_n, column=ci, value=v)
            _sc(cell, bg=COLOR["diferencia"], fc=COLOR["sin_font"],
                fmt=fmt, ha="right" if ci in [4,5,6] else "center" if ci in [1,7] else "left")
        ws.row_dimensions[row_n].height = 16
        row_n += 1

    # Subtotal A
    for ci,v in enumerate(["","","SUBTOTAL",total_aux_pdte,"","","",""],1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        _sc(cell, bg=COLOR["total_bg"], fc="FFFFFF", bold=True,
            fmt=FMT_COP if ci==4 else None,
            ha="right" if ci==4 else "left" if ci==3 else "center")
    ws.row_dimensions[row_n].height = 18
    row_n += 2

    # ── B. Datafono sin contrapartida en auxiliar ────────────────────────────
    # Días que existen en el datafono pero no en ningún registro del auxiliar
    dias_en_aux = set()
    for res in resultados:
        d = res['row_aux']['Dia_Operacion']
        if pd.notna(d):
            dias_en_aux.add(int(d))

    df_sin_aux = por_dia[
        (por_dia['_Nombre_Archivo'].str.upper() == nombre_archivo.upper()) &
        (por_dia['_Year_Vale'] == year) &
        (por_dia['_Mes_Vale']  == mes)  &
        (~por_dia['_Dia_Vale'].isin(dias_en_aux))
    ].copy()

    ws.merge_cells(f"A{row_n}:J{row_n}")
    h2 = ws.cell(row=row_n, column=1,
                 value=f"B)  EN DATAFONO SIN REGISTRO EN AUXILIAR ({len(df_sin_aux)} días)")
    h2.fill = PatternFill("solid", start_color=COLOR["dif_menor"])
    h2.font = Font(color=COLOR["dif_font"], bold=True, name="Arial", size=10)
    h2.alignment = Alignment(horizontal="left")
    ws.row_dimensions[row_n].height = 20
    row_n += 1

    HEADERS_B = ["Día Fecha Vale","Valor Neto Datafono ($)","Comisión + Ret ($)","Observación"]
    widths_b   = [16, 24, 22, 44]
    for ci,(h,w) in enumerate(zip(HEADERS_B,widths_b),1):
        cell = ws.cell(row=row_n, column=ci, value=h)
        _sc(cell, bg=COLOR["header_bg"], fc=COLOR["header_font"], bold=True, ha="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(
            ws.column_dimensions[get_column_letter(ci)].width, w)
    ws.row_dimensions[row_n].height = 18
    row_n += 1

    total_df_pdte = 0
    for _, gr in df_sin_aux.iterrows():
        com_ret = sum(float(gr[c]) for c in
                      ['Valor Comisión','Ret. Fuente','Ret. IVA','Ret. ICA']
                      if c in gr.index and pd.notna(gr[c]))
        v_neto = gr['Valor Neto']
        total_df_pdte += v_neto
        vals = [int(gr['_Dia_Vale']), v_neto,
                com_ret if com_ret > 0 else None,
                "Abono recibido en banco sin registro contable en el auxiliar"]
        fmts = [None, FMT_COP, FMT_COP, None]
        for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
            cell = ws.cell(row=row_n, column=ci, value=v)
            _sc(cell, bg=COLOR["dif_menor"], fc=COLOR["dif_font"],
                fmt=fmt, ha="right" if ci in [2,3] else "center" if ci==1 else "left")
        ws.row_dimensions[row_n].height = 16
        row_n += 1

    # Subtotal B
    for ci,v in enumerate(["SUBTOTAL",total_df_pdte,"",""],1):
        cell = ws.cell(row=row_n, column=ci, value=v)
        _sc(cell, bg=COLOR["total_bg"], fc="FFFFFF", bold=True,
            fmt=FMT_COP if ci==2 else None,
            ha="right" if ci==2 else "left")
    ws.row_dimensions[row_n].height = 18
    row_n += 2

    # ── Resumen al pie ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{row_n}:J{row_n}")
    res_cell = ws.cell(row=row_n, column=1)
    res_cell.value = (
        f"RESUMEN PENDIENTES:  "
        f"Auxiliar sin conciliar = ${total_aux_pdte:,.0f}  |  "
        f"Datafono sin registrar = ${total_df_pdte:,.0f}  |  "
        f"Diferencia neta pendiente = ${total_aux_pdte - total_df_pdte:,.0f}"
    )
    res_cell.fill = PatternFill("solid", start_color="1F3864")
    res_cell.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    res_cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[row_n].height = 22
    ws.freeze_panes = "A5"

def generar_excel_resultado(resultados: list, sede_nombre: str,
                             output_path: str, periodo: str = "Enero 2026",
                             info_match: dict = None,
                             por_dia=None, por_bolruta=None,
                             nombre_archivo: str = "",
                             year: int = 2026, mes: int = 1,
                             modo_fecha: str = 'FECHA_VALE') -> str:
    """Genera Excel individual para una sede."""
    wb = Workbook()
    _hoja_resumen_individual(wb, resultados, sede_nombre, periodo, info_match)
    _hoja_detalle(wb, resultados, sede_nombre, periodo)
    if por_dia is not None and nombre_archivo:
        _hoja_pendientes(wb, resultados, por_dia, por_bolruta,
                         nombre_archivo, year, mes, sede_nombre)
    _hoja_auditoria(wb, sede_nombre, periodo, info_match, modo_fecha=modo_fecha)
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    wb.save(output_path)
    return output_path


def generar_excel_unificado(resultados_por_sede: list,
                             output_path: str,
                             periodo: str,
                             mapa_nombres: dict,
                             huerfanos: list,
                             modo_fecha: str = 'FECHA_VALE') -> str:
    """
    Genera un único Excel con todas las sedes.
    resultados_por_sede: [{'sede', 'nombre_archivo', 'resultados', 'info_match'}]
    """
    wb = Workbook()

    resumen_sedes = []
    for item in resultados_por_sede:
        res = item['resultados']
        resumen_sedes.append({
            'sede':       item['sede'],
            'nombre_df':  item['nombre_archivo'],
            'total_reg':  len(res),
            'cuadra':     sum(1 for r in res if r['estado']=='CUADRA'),
            'dif_menor':  sum(1 for r in res if r['estado']=='DIF_MENOR'),
            'revisar':    sum(1 for r in res if r['estado'] in ('DIFERENCIA','SIN_MATCH','SIN_DIA')),
            'sin_df':     sum(1 for r in res if r['estado']=='SIN_MATCH'),
            'total_aux':  sum(r['valor_aux']   for r in res),
            'total_df':   sum(r['suma_grupos'] for r in res),
            'diferencia': sum(r['valor_aux']   for r in res) - sum(r['suma_grupos'] for r in res),
            'score':      item.get('info_match',{}).get('score',0),
            'tipo_match': item.get('info_match',{}).get('tipo','—'),
        })

    _hoja_resumen_ejecutivo(wb, resumen_sedes, huerfanos, periodo)
    _hoja_mapa_nombres(wb, mapa_nombres, huerfanos)

    for item in resultados_por_sede:
        nombre_hoja = item['sede'][:28].strip()
        _hoja_detalle(wb, item['resultados'], item['sede'], periodo, ws_name=nombre_hoja)
        _hoja_pendientes(wb, item['resultados'],
                         item.get('por_dia', pd.DataFrame()),
                         item.get('por_bolruta', pd.DataFrame()),
                         item['nombre_archivo'],
                         item.get('year', 2026), item.get('mes', 1),
                         item['sede'])

    _hoja_auditoria(wb, "Consolidado", periodo, es_consolidado=True,
                    mapa_nombres=mapa_nombres, huerfanos=huerfanos,
                    modo_fecha=modo_fecha)

    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    wb.save(output_path)
    return output_path


def generar_resumen_consolidado(resumen_sedes: list, output_path: str,
                                 periodo: str, huerfanos: list = None) -> str:
    """Genera solo el resumen ejecutivo (modo archivos separados)."""
    wb = Workbook()
    _hoja_resumen_ejecutivo(wb, resumen_sedes, huerfanos or [], periodo)
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    wb.save(output_path)
    return output_path
